# ----------------------------
# Enforce Administrator Privileges on Windows
# ----------------------------
import ctypes
import sys
import os
from collections import deque
import winsound  # For sound playback
def is_admin():
    """Check if the script is running with administrator privileges."""
    try:
        return ctypes.windll.shell32.IsUserAnAdmin()
    except AttributeError:
        return True
    except:
        return False
if not is_admin():
    ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, " ".join(sys.argv), None, 1)
    sys.exit()

# ----------------------------
# Centralized App Data Directory
# ----------------------------
APP_DATA_DIR = os.path.join(os.environ.get("PROGRAMDATA", "C:\\ProgramData"), "Batching Recipe")
ASSETS_DIR = os.path.join(APP_DATA_DIR, "assets")
os.makedirs(ASSETS_DIR, exist_ok=True)
ALARM_SOUND_FILE = os.path.join(ASSETS_DIR, "alarm.wav")

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import serial  # type: ignore[import]
import serial.tools.list_ports  # type: ignore[import]
import threading
import time
import re
import csv
import json
import logging
from datetime import datetime
from logging.handlers import RotatingFileHandler
import subprocess
from openpyxl import Workbook  # type: ignore[import]
from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore[import]

# ----------------------------
# Logging configuration
# ----------------------------
LOG_FILE = os.path.join(APP_DATA_DIR, "batch_app.log")
log_handler = RotatingFileHandler(LOG_FILE, maxBytes=10*1024*1024, backupCount=5)
log_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
logging.basicConfig(level=logging.INFO, handlers=[log_handler, logging.StreamHandler()])

# ----------------------------
# Constants and helpers
# ----------------------------
COLOR_UNDER = "#FFC107"
COLOR_OK = "#4CAF50"
COLOR_OVER = "#F44336"
COLOR_TEXT_DEFAULT = "#222222"
CSV_FILE = os.path.join(APP_DATA_DIR, "batch_transactions.csv")
RECIPES_DB_FILE = os.path.join(APP_DATA_DIR, "recipes.json")
BATCH_STATES_FILE = os.path.join(APP_DATA_DIR, "batch_states.json")
SETTINGS_FILE = os.path.join(APP_DATA_DIR, "user_settings.json")
CORRECT_PASSWORD = "789123"
MAX_PASSWORD_ATTEMPTS = 3

# ESC/POS Commands
ESC = b'\x1b'
GS = b'\x1d'
LF = b'\n'
FULL_CUT = GS + b'V\x00'          # Full cut
PARTIAL_CUT = GS + b'V\x01'       # Partial cut (if supported)
BOLD_ON = ESC + b'E\x01'
BOLD_OFF = ESC + b'E\x00'
ALIGN_LEFT = ESC + b'a\x00'
ALIGN_CENTER = ESC + b'a\x01'
ALIGN_RIGHT = ESC + b'a\x02'
FEED_PAPER = ESC + b'd\x03'       # Feed 3 lines

def ensure_csv_header(path=CSV_FILE):
    header = [
        "batch_id", "transaction_number", "timestamp", "ingredient_index", "ingredient_name",
        "target_weight_g", "tolerance_percent", "min_ok_g", "max_ok_g",
        "actual_weight_g", "status", "deviation_g", "units"
    ]
    need_header = not os.path.exists(path) or os.path.getsize(path) == 0
    if need_header:
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(header)

def format_float(x, nd=2):
    try:
        return f"{float(x):.{nd}f}"
    except Exception:
        return str(x)

def prompt_password_with_retry(parent, title, prompt):
    for attempt in range(MAX_PASSWORD_ATTEMPTS):
        password = simpledialog.askstring(
            title,
            f"{prompt}\nAttempt {attempt + 1} of {MAX_PASSWORD_ATTEMPTS}:",
            show='*',
            parent=parent
        )
        if password == CORRECT_PASSWORD:
            return True
        elif password is None:
            return False
        else:
            messagebox.showerror("Access Denied", f"Incorrect password. {MAX_PASSWORD_ATTEMPTS - attempt - 1} attempts remaining.")
    messagebox.showerror("Access Denied", "Maximum password attempts exceeded. Action canceled.")
    return False

# ----------------------------
# Main Application
# ----------------------------
class BatchingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Batching Recipes - Scale/Serial")
        try:
            self.root.iconbitmap(os.path.join(ASSETS_DIR, "icon.ico"))
        except:
            pass
        self.root.geometry("1400x650")
        self.root.rowconfigure(0, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.main_container = tk.Frame(self.root)
        self.main_container.grid(row=0, column=0, sticky="nsew")
        self.main_container.rowconfigure(0, weight=1)
        self.main_container.columnconfigure(0, weight=1)
        self.connect_btn = None
        self.disconnect_btn = None
        self.alarm_connect_btn = None
        self.alarm_disconnect_btn = None
        self.com_combo = None
        self.baud_combo = None
        self.parity_combo = None
        self.alarm_com_combo = None
        self.alarm_baud_combo = None
        self.alarm_parity_combo = None
        self.recipe_combo = None
        self.trans_tree = None
        self.trans_window = None
        self.developer_frame = None
        self.developer_visible = False
        self.start_btn = None
        self.last_serial_line = ""
        self.raw_serial_lines = deque(maxlen=4)
        self.predefined_regexes = [
            r"wn(\d+\.?\d*)kg",
            r"ww(-?\d+)",
            r"(-?\d+\.\d+)\s*kg",
            r"ST,GS,(\d+\.\d+)",
            r"W=(\d+\.\d+)",
            r"([\+\-]?\d+\.\d+)",
            r"(\d+\.\d+)",
            r"(\d+)",
            r"([+-]?\d*\.?\d+)",
            "Custom"
        ]
        self.data_format_regex_var = tk.StringVar(value=r"(\d+\.?\d*)")
        self.regex_fail_count = 0
        self.load_user_settings()
        self.ser = None
        self.alarm_ser = None
        self.read_thread = None
        self.alarm_read_thread = None
        self.stop_read_event = threading.Event()
        self.stop_alarm_event = threading.Event()
        self.connected = False
        self.alarm_connected = False
        self.weight_lock = threading.Lock()
        self._current_weight = 0.0
        self.current_status = "UNDER"
        self.flash_job = None
        self.flash_on = False
        self.flash_color = COLOR_UNDER
        self.auto_advance_enabled = tk.BooleanVar(value=self.settings.get("auto_advance_enabled", True))
        self.hold_time_ms = tk.IntVar(value=self.settings.get("hold_time_ms", 1000))
        self.in_tolerance_since = None
        self.last_manual_capture = 0
        self.waiting_for_zero = False
        self.zero_stable_since = None
        self.zero_tolerance = 0.5
        self.tolerance_percent = tk.DoubleVar(value=self.settings.get("tolerance_percent", 2.0))
        self.units = tk.StringVar(value=self.settings.get("units", "g"))
        self.simulate_mode = tk.BooleanVar(value=self.settings.get("simulate_mode", True))
        self.sim_weight_var = tk.DoubleVar(value=0.0)
        self.recipe = []
        self.current_recipe_name = None
        self.recipes_db = {}
        self.current_index = -1
        self.batch_active = False
        self.batch_id = None
        self.net_total = 0.0
        self.grand_total_net = 0.0
        self.batch_states = {}
        self.user_batch_id_var = tk.StringVar(value=self.settings.get("user_batch_id", ""))
        self.lock_batch_id = tk.BooleanVar(value=self.settings.get("lock_batch_id", False))
        self.auto_loop_enabled = tk.BooleanVar(value=self.settings.get("auto_loop_enabled", False))
        self.alarm_send_under = tk.BooleanVar(value=self.settings.get("alarm_send_under", True))
        self.alarm_send_ok = tk.BooleanVar(value=self.settings.get("alarm_send_ok", False))
        self.alarm_send_over = tk.BooleanVar(value=self.settings.get("alarm_send_over", True))
        self.alarm_sound_under = tk.BooleanVar(value=self.settings.get("alarm_sound_under", False))
        self.alarm_sound_ok = tk.BooleanVar(value=self.settings.get("alarm_sound_ok", False))
        self.alarm_sound_over = tk.BooleanVar(value=self.settings.get("alarm_sound_over", False))
        self.alarm_interval_ms = tk.IntVar(value=self.settings.get("alarm_interval_ms", 100))
        self.alarm_sound_interval_ms = tk.IntVar(value=self.settings.get("alarm_sound_interval_ms", 500))
        self.alarm_com_var = tk.StringVar(value=self.settings.get("alarm_com", self.get_default_com()))
        self.alarm_baud_var = tk.StringVar(value=self.settings.get("alarm_baud", "9600"))
        self.alarm_parity_var = tk.StringVar(value=self.settings.get("alarm_parity", "None"))
        self.alarm_data_bits_var = tk.StringVar(value=self.settings.get("alarm_data_bits", "8"))
        self.alarm_stop_bits_var = tk.StringVar(value=self.settings.get("alarm_stop_bits", "1"))
        self.com_var = tk.StringVar(value=self.settings.get("com", self.get_default_com()))
        self.baud_var = tk.StringVar(value=self.settings.get("baud", "9600"))
        self.parity_var = tk.StringVar(value=self.settings.get("parity", "None"))
        self.data_bits_var = tk.StringVar(value=self.settings.get("data_bits", "8"))
        self.stop_bits_var = tk.StringVar(value=self.settings.get("stop_bits", "1"))
        self.read_interval_us = tk.IntVar(value=self.settings.get("read_interval_us", 100000))
        self.auto_detect_threshold = tk.IntVar(value=self.settings.get("auto_detect_threshold", 3))
        self.scale_units_var = tk.StringVar(value=self.settings.get("scale_units", "kg"))
        self.lock_stable_var = tk.BooleanVar(value=self.settings.get("lock_stable_reading", False))
        self.stable_threshold_var = tk.IntVar(value=self.settings.get("stable_threshold", 2))
        self.decimal_places_var = tk.IntVar(value=self.settings.get("decimal_places", 2))
        self.printer_ser = None
        self.printer_name = None  # For default printer mode
        self.printer_connected = False
        self.print_paper_width = tk.IntVar(value=self.settings.get("print_paper_width", 32))
        # Trace callbacks for auto-save
        for var in [
            self.auto_advance_enabled, self.hold_time_ms, self.tolerance_percent, self.units,
            self.simulate_mode, self.user_batch_id_var, self.lock_batch_id, self.auto_loop_enabled,
            self.alarm_send_under, self.alarm_send_ok, self.alarm_send_over,
            self.alarm_sound_under, self.alarm_sound_ok, self.alarm_sound_over,
            self.alarm_interval_ms, self.alarm_sound_interval_ms,
            self.alarm_com_var, self.alarm_baud_var, self.alarm_parity_var,
            self.alarm_data_bits_var, self.alarm_stop_bits_var,
            self.com_var, self.baud_var, self.parity_var, self.data_bits_var, self.stop_bits_var,
            self.read_interval_us, self.auto_detect_threshold, self.scale_units_var,
            self.lock_stable_var, self.stable_threshold_var, self.decimal_places_var,
            self.data_format_regex_var, self.enable_print_var, self.printer_mode_var, self.printer_com_var, 
            self.printer_baud_var, self.printer_parity_var, self.print_header_var, self.print_paper_width
        ]:
            var.trace_add("write", lambda *a: self.save_user_settings())
        self.validate_paper_width()
        self.build_ui()
        ensure_csv_header()
        self.load_recipes_db()
        self.load_batch_states()
        self.root.after(120, self.ui_tick)
        self.on_sim_toggle()
        # ✅ Keyboard shortcuts
        self.root.bind("<F2>", self.on_edit_shortcut)
        self.root.bind("<Delete>", self.on_delete_shortcut)
        self.root.bind("<Control-n>", self.on_new_recipe_shortcut)
        self.root.bind("<Control-r>", self.on_resume_shortcut)

    def validate_paper_width(self, *args):
        try:
            val = int(self.print_paper_width.get())
            self.print_paper_width.set(max(24, min(48, val)))
        except ValueError:
            self.print_paper_width.set(32)

    # =================== EDITING METHODS (MODAL) ===================
    def on_tree_double_click(self, event):
        item_id = self.tree.focus()
        if not item_id:
            return
        idx = self.tree.index(item_id)
        self.edit_ingredient_at_index(idx)

    def edit_selected(self):
        if self.tree is None:
            return
        selection = self.tree.selection()
        if not selection:
            messagebox.showinfo("Info", "Please select an ingredient to edit.")
            return
        if len(selection) > 1:
            messagebox.showinfo("Info", "Please select only one ingredient to edit.")
            return
        if len(selection) >= 1:
            item_id = selection[0]
            idx = self.tree.index(item_id)
            self.edit_ingredient_at_index(idx)

    def edit_ingredient_at_index(self, idx):
        if idx < 0 or idx >= len(self.recipe):
            return
        item = self.recipe[idx]
        old_name = item["name"]
        old_target = item["target"]
        old_tol = item["tol"]
        new_name = simpledialog.askstring("Edit Ingredient", "Ingredient Name:", initialvalue=old_name)
        if new_name is None:
            return
        new_name = new_name.strip()
        if not new_name:
            messagebox.showerror("Invalid Input", "Name cannot be empty.")
            return
        # ✅ Prevent duplicate names in same recipe
        new_name_lower = new_name.lower()
        for i, ing in enumerate(self.recipe):
            if i != idx and ing["name"].lower() == new_name_lower:
                messagebox.showerror("Duplicate Name", "An ingredient with this name already exists in the recipe.")
                return
        new_target_str = simpledialog.askstring("Edit Ingredient", "Target Weight (g):", initialvalue=str(old_target))
        if new_target_str is None:
            return
        try:
            new_target = float(new_target_str)
        except (ValueError, TypeError):
            messagebox.showerror("Invalid Input", "Target must be a number.")
            return
        tol_str = "" if old_tol is None else str(old_tol)
        new_tol_str = simpledialog.askstring("Edit Ingredient", "Tolerance (%): (leave blank for default)", initialvalue=tol_str)
        if new_tol_str is None:
            return
        new_tol = None
        if new_tol_str.strip() != "":
            try:
                new_tol = float(new_tol_str)
                if new_tol < 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("Invalid Input", "Tolerance must be non-negative or blank.")
                return
        self.recipe[idx]["name"] = new_name
        self.recipe[idx]["target"] = new_target
        self.recipe[idx]["tol"] = new_tol
        tol_disp = "-" if new_tol is None else format_float(new_tol)
        if self.tree is not None:
            children = self.tree.get_children()
            if idx < len(children):
                self.tree.item(
                    children[idx],
                    values=(new_name, format_float(new_target), tol_disp, self.recipe[idx]["actual"] or "-", self.recipe[idx]["status"])
                )
        self.save_recipes_db()
        self.log(f"Edited ingredient: {new_name}")

    def on_edit_shortcut(self, event=None):
        self.on_tree_double_click(event)

    def on_delete_shortcut(self, event=None):
        self.delete_selected()

    def on_new_recipe_shortcut(self, event=None):
        self.create_new_recipe()

    def on_resume_shortcut(self, event=None):
        batch_id = simpledialog.askstring("Resume Batch", "Enter Batch ID to resume:")
        if batch_id:
            self.resume_batch_id_var.set(batch_id)
            self.resume_batch_by_id()

    # =================== PRINTER METHODS ===================
    def load_user_settings(self):
        if os.path.exists(SETTINGS_FILE):
            try:
                with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                    self.settings = json.load(f)
                logging.info(f"User settings loaded from {SETTINGS_FILE}")
            except Exception as e:
                logging.error(f"Failed to load settings: {e}")
                self.settings = {}
        else:
            self.settings = {}
        # Printer settings
        self.enable_print_var = tk.BooleanVar(value=self.settings.get("enable_auto_print", False))
        self.printer_mode_var = tk.StringVar(value=self.settings.get("printer_mode", "COM"))  # "COM" or "DEFAULT"
        self.printer_com_var = tk.StringVar(value=self.settings.get("printer_com", "COM3"))
        self.printer_baud_var = tk.StringVar(value=self.settings.get("printer_baud", "9600"))
        self.printer_parity_var = tk.StringVar(value=self.settings.get("printer_parity", "None"))
        self.print_header_var = tk.StringVar(value=self.settings.get("print_header", "COMPANY NAME\nAddr • Tel"))

    def save_user_settings(self):
        settings_to_save = {
            "auto_advance_enabled": self.auto_advance_enabled.get(),
            "hold_time_ms": self.hold_time_ms.get(),
            "tolerance_percent": self.tolerance_percent.get(),
            "units": self.units.get(),
            "simulate_mode": self.simulate_mode.get(),
            "com": self.com_var.get(),
            "baud": self.baud_var.get(),
            "parity": self.parity_var.get(),
            "data_bits": self.data_bits_var.get(),
            "stop_bits": self.stop_bits_var.get(),
            "alarm_com": self.alarm_com_var.get(),
            "alarm_baud": self.alarm_baud_var.get(),
            "alarm_parity": self.alarm_parity_var.get(),
            "alarm_data_bits": self.alarm_data_bits_var.get(),
            "alarm_stop_bits": self.alarm_stop_bits_var.get(),
            "alarm_send_under": self.alarm_send_under.get(),
            "alarm_send_ok": self.alarm_send_ok.get(),
            "alarm_send_over": self.alarm_send_over.get(),
            "alarm_sound_under": self.alarm_sound_under.get(),
            "alarm_sound_ok": self.alarm_sound_ok.get(),
            "alarm_sound_over": self.alarm_sound_over.get(),
            "alarm_interval_ms": self.alarm_interval_ms.get(),
            "alarm_sound_interval_ms": self.alarm_sound_interval_ms.get(),
            "user_batch_id": self.user_batch_id_var.get(),
            "lock_batch_id": self.lock_batch_id.get(),
            "auto_loop_enabled": self.auto_loop_enabled.get(),
            "input_data_regex": self.data_format_regex_var.get(),
            "read_interval_us": self.read_interval_us.get(),
            "auto_detect_threshold": self.auto_detect_threshold.get(),
            "scale_units": self.scale_units_var.get(),
            "lock_stable_reading": self.lock_stable_var.get(),
            "stable_threshold": self.stable_threshold_var.get(),
            "decimal_places": self.decimal_places_var.get(),
            "enable_auto_print": self.enable_print_var.get(),
            "printer_mode": self.printer_mode_var.get(),
            "printer_com": self.printer_com_var.get(),
            "printer_baud": self.printer_baud_var.get(),
            "printer_parity": self.printer_parity_var.get(),
            "print_header": self.print_header_var.get(),
            "print_paper_width": self.print_paper_width.get(),
        }
        try:
            with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
                json.dump(settings_to_save, f, indent=2)
            logging.info(f"User settings saved to {SETTINGS_FILE}")
        except Exception as e:
            logging.error(f"Failed to save settings: {e}")

    def on_printer_mode_change(self):
        """Toggle visibility of COM port controls based on printer mode."""
        if self.printer_mode_var.get() == "COM":
            self.com_controls_frame.pack(side=tk.LEFT, padx=(0, 0))
        else:
            self.com_controls_frame.pack_forget()
        self.save_user_settings()

    def on_print_toggle(self):
        if self.enable_print_var.get():
            self.connect_printer()
        else:
            self.disconnect_printer()

    def connect_printer(self):
        if self.printer_connected:
            return
        
        mode = self.printer_mode_var.get()
        
        if mode == "DEFAULT":
            # Use Windows default printer
            try:
                import win32print  # type: ignore[import]
                self.printer_name = win32print.GetDefaultPrinter()
                self.printer_connected = True
                self.printer_ser = None  # Not used for default printer
                self.update_printer_status_label()
                self.log(f"Printer connected: Default printer '{self.printer_name}'")
            except ImportError:
                # win32print not available, show error
                self.printer_connected = False
                self.enable_print_var.set(False)
                self.update_printer_status_label()
                messagebox.showerror("Printer Error", "Default printer mode requires win32print package.\nPlease use COM port mode or install pywin32.")
                self.log("Failed to connect: win32print not installed", level="ERROR")
            except Exception as e:
                self.printer_connected = False
                self.enable_print_var.set(False)
                self.update_printer_status_label()
                messagebox.showerror("Printer Error", f"Failed to connect to default printer:\n{str(e)}")
                self.log(f"Default printer connection failed: {e}", level="ERROR")
        else:
            # Use COM port
            parity_map = {"None": serial.PARITY_NONE, "Even": serial.PARITY_EVEN, "Odd": serial.PARITY_ODD}
            try:
                self.printer_ser = serial.Serial(
                    port=self.printer_com_var.get(),
                    baudrate=int(self.printer_baud_var.get()),
                    parity=parity_map.get(self.printer_parity_var.get(), serial.PARITY_NONE),
                    bytesize=8,
                    stopbits=serial.STOPBITS_ONE,
                    timeout=1,
                    write_timeout=2
                )
                self.printer_connected = True
                self.printer_name = None  # Not used for COM port
                self.update_printer_status_label()
                self.log(f"Printer connected: {self.printer_com_var.get()} @ {self.printer_baud_var.get()} baud, parity={self.printer_parity_var.get()}")
            except Exception as e:
                self.printer_connected = False
                if self.printer_ser:
                    try:
                        self.printer_ser.close()
                    except:
                        pass
                    self.printer_ser = None
                self.enable_print_var.set(False)
                self.update_printer_status_label()
                messagebox.showerror("Printer Error", f"Failed to connect printer:\n{str(e)}")
                self.log(f"Printer connection failed: {e}", level="ERROR")

    def disconnect_printer(self):
        if self.printer_ser:
            try:
                self.printer_ser.close()
            except:
                pass
            self.printer_ser = None
        self.printer_connected = False
        self.update_printer_status_label()
        self.log("Printer disconnected.")

    def update_printer_status_label(self):
        if not self.printer_connected:
            self.printer_status_label.config(text="Printer: Disconnected", fg="#9E9E9E")
        else:
            try:
                if self.printer_ser and self.printer_ser.is_open:
                    self.printer_status_label.config(text="Printer: OK", fg=COLOR_OK)
                else:
                    self.printer_status_label.config(text="Printer: Error", fg=COLOR_OVER)
            except:
                self.printer_status_label.config(text="Printer: Error", fg=COLOR_OVER)

    def _send_print_job(self, data):
        """Runs in background thread to avoid UI freeze."""
        try:
            if self.printer_mode_var.get() == "DEFAULT":
                self._print_transaction_ticket_default(data)
            else:
                self._print_transaction_ticket_sync(data)
            self.log("Ticket printed.")
        except Exception as e:
            self.log(f"Print error: {e}", level="ERROR")
            self.root.after(0, lambda: messagebox.showwarning("Print Failed", "Failed to send ticket to printer."))

    def _print_transaction_ticket_default(self, data):
        """Print to Windows default printer."""
        try:
            import win32print  # type: ignore[import]
            import win32api  # type: ignore[import]
            
            ticket_text = self.format_ticket_text(data)
            
            # Use Notepad to print (generic approach)
            import tempfile
            with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False) as f:
                f.write(ticket_text)
                temp_file = f.name
            
            try:
                # Print the file using Windows default printer
                win32api.ShellExecute(0, "print", temp_file, "", ".", 0)
                time.sleep(1)  # Give printer time to queue
            finally:
                # Clean up temp file
                try:
                    os.remove(temp_file)
                except:
                    pass
        except ImportError:
            self.log("win32print not available for default printer", level="ERROR")
        except Exception as e:
            self.log(f"Default printer error: {e}", level="ERROR")

    def _print_transaction_ticket_sync(self, data):
        """Synchronous print logic for COM port (called from thread)."""
        if not self.enable_print_var.get() or not self.printer_connected or self.printer_ser is None:
            return
        width = self.print_paper_width.get()
        separator = b'-' * width + LF
        footer_line = b'=' * width + LF
        buf = b''
        buf += ALIGN_CENTER
        buf += BOLD_ON
        for line in self.print_header_var.get().split('\n'):
            buf += line.encode('ascii', errors='replace') + LF
        buf += BOLD_OFF
        buf += separator
        buf += ALIGN_LEFT
        buf += f"Batch: {data['batch_id']}".encode() + LF
        buf += f"Time:  {data['timestamp']}".encode() + LF
        buf += f"Item:  {data['ingredient']}".encode() + LF
        buf += f"Target: {data['target']:.2f} g".encode() + LF
        buf += f"Actual: {data['actual']:.2f} g".encode() + LF
        buf += f"Tol:    {data['tolerance']:.1f}%".encode() + LF
        buf += f"Status: {data['status']}".encode() + LF
        buf += f"Dev:    {data['deviation']:+.2f} g".encode() + LF
        buf += footer_line
        buf += FEED_PAPER
        buf += FULL_CUT
        self.printer_ser.write(buf)
        self.printer_ser.flush()

    def print_transaction_ticket(self, data):
        """Public method — starts background print job."""
        if not self.enable_print_var.get() or not self.printer_connected:
            return
        threading.Thread(target=self._send_print_job, args=(data,), daemon=True).start()

    def format_ticket_text(self, data):
        """Generate the text representation of the ticket for preview and printing."""
        width = self.print_paper_width.get()
        lines = []
        
        # Header
        lines.append("=" * width)
        for header_line in self.print_header_var.get().split('\n'):
            lines.append(header_line.center(width))
        lines.append("=" * width)
        
        # Transaction details
        lines.append("")
        lines.append(f"Batch: {data['batch_id']}")
        lines.append(f"Time:  {data['timestamp']}")
        lines.append("")
        
        # Ingredient and weight
        lines.append(f"Item:  {data['ingredient']}")
        lines.append("")
        lines.append(f"Target: {data['target']:.2f} g")
        lines.append(f"Actual: {data['actual']:.2f} g")
        lines.append(f"Tol:    {data['tolerance']:.1f}%")
        lines.append("")
        
        # Status and deviation
        lines.append(f"Status: {data['status']}")
        lines.append(f"Dev:    {data['deviation']:+.2f} g")
        lines.append("")
        lines.append("=" * width)
        
        return "\n".join(lines)

    def show_print_preview(self, data):
        """Display a preview window of the ticket to be printed."""
        preview_win = tk.Toplevel(self.root)
        preview_win.title("Print Ticket Preview")
        preview_win.geometry("450x600")
        preview_win.resizable(False, False)
        
        # Center the preview window
        preview_win.transient(self.root)
        preview_win.grab_set()
        
        # Title label
        title_label = tk.Label(
            preview_win,
            text="Print Preview - Thermal Receipt (32 char width)",
            font=("Arial", 10, "bold"),
            bg="#f0f0f0",
            pady=5
        )
        title_label.pack(fill=tk.X)
        
        # Preview frame with scrollbar
        frame = tk.Frame(preview_win)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Text widget for preview
        text_widget = tk.Text(
            frame,
            font=("Courier New", 10),
            height=20,
            width=38,
            wrap=tk.WORD,
            relief="sunken",
            borderwidth=2,
            bg="white",
            fg="black"
        )
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Scrollbar
        scrollbar = tk.Scrollbar(frame, command=text_widget.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        text_widget.config(yscrollcommand=scrollbar.set)
        
        # Insert formatted text
        ticket_text = self.format_ticket_text(data)
        text_widget.insert(tk.END, ticket_text)
        text_widget.config(state="disabled")
        
        # Button frame
        btn_frame = tk.Frame(preview_win)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)
        btn_frame.columnconfigure(0, weight=1)
        btn_frame.columnconfigure(1, weight=1)
        
        # Print button
        print_btn = tk.Button(
            btn_frame,
            text="🖨️ Print Ticket",
            bg="#4CAF50",
            fg="white",
            font=("Arial", 10, "bold"),
            command=lambda: self._print_and_close_preview(data, preview_win)
        )
        print_btn.grid(row=0, column=0, sticky="ew", padx=(0, 5))
        
        # Copy to clipboard button
        copy_btn = tk.Button(
            btn_frame,
            text="📋 Copy Text",
            bg="#2196F3",
            fg="white",
            font=("Arial", 10, "bold"),
            command=lambda: self._copy_preview_text(ticket_text, preview_win)
        )
        copy_btn.grid(row=0, column=1, sticky="ew", padx=(5, 0))
        
        # Close button
        close_btn = tk.Button(
            btn_frame,
            text="Close",
            bg="#9E9E9E",
            fg="white",
            font=("Arial", 10),
            command=preview_win.destroy
        )
        close_btn.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(5, 0))

    def _print_and_close_preview(self, data, preview_window):
        """Send ticket to printer and close preview window."""
        if not self.enable_print_var.get():
            messagebox.showwarning("Printer Disabled", "Printer is not enabled.")
            return
        if not self.printer_connected:
            messagebox.showwarning("Printer Disconnected", "Printer is not connected. Connect printer and try again.")
            return
        threading.Thread(target=self._send_print_job, args=(data,), daemon=True).start()
        messagebox.showinfo("Print Job Sent", "Ticket sent to printer.")
        preview_window.destroy()

    def _copy_preview_text(self, text, preview_window):
        """Copy preview text to clipboard."""
        try:
            preview_window.clipboard_clear()
            preview_window.clipboard_append(text)
            messagebox.showinfo("Copied", "Ticket text copied to clipboard.")
        except Exception as e:
            messagebox.showerror("Copy Error", f"Failed to copy: {e}")

    def test_print_ticket(self):
        """Show print preview with test data."""
        test_data = {
            "batch_id": "TEST-001",
            "timestamp": datetime.now().strftime("%m/%d/%Y %I:%M:%S %p"),
            "ingredient": "Test Ingredient",
            "target": 1000.0,
            "actual": 1005.3,
            "tolerance": 2.0,
            "status": "OK",
            "deviation": +5.3
        }
        self.show_print_preview(test_data)
        self.log("Print preview window opened.")

    # =================== CORE ===================
    def get_consistent_batch_id(self):
        if self.batch_id:
            return self.batch_id
        if self.lock_batch_id.get():
            user_input = self.user_batch_id_var.get().strip()
            if user_input:
                return user_input
            batch_state = self.batch_states.get(self.current_recipe_name, {})
            locked_id = batch_state.get("locked_batch_id")
            if not locked_id:
                locked_id = f"{self.current_recipe_name}_{datetime.now().strftime('%H%M%S')}"
                batch_state["locked_batch_id"] = locked_id
                self.batch_states[self.current_recipe_name] = batch_state
                self.save_batch_states()
            return locked_id
        else:
            return f"{self.current_recipe_name}_{datetime.now().strftime('%H%M%S')}"

    def on_sim_toggle(self):
        sim = self.simulate_mode.get()
        if self.connect_btn:
            self.connect_btn.configure(state="normal" if not sim else "disabled")
        if self.disconnect_btn:
            self.disconnect_btn.configure(state="normal" if self.connected else "disabled")
        serial_state = "disabled" if sim else "normal"
        for w in [self.com_combo, self.baud_combo, self.parity_combo, self.data_format_regex_combo]:
            if w:
                w.configure(state=serial_state)
        if self.alarm_connect_btn:
            self.alarm_connect_btn.configure(state="normal" if not self.alarm_connected else "disabled")
        if self.alarm_disconnect_btn:
            self.alarm_disconnect_btn.configure(state="normal" if self.alarm_connected else "disabled")
        for w in [self.alarm_com_combo, self.alarm_baud_combo, self.alarm_parity_combo]:
            if w:
                w.configure(state="normal")

    def toggle_developer_panels(self):
        if not prompt_password_with_retry(self.root, "Password Required", "Enter password to toggle Developer Panels:"):
            return
        if self.developer_visible:
            if self.developer_frame is not None:
                self.developer_frame.pack_forget()
            self.settings_menu.entryconfig(0, label="Show Developer Panels")
            self.developer_visible = False
        else:
            if self.developer_frame is not None:
                self.developer_frame.pack(fill=tk.X, before=self.main_frame)
            self.settings_menu.entryconfig(0, label="Hide Developer Panels")
            self.developer_visible = True

    def validate_read_interval(self, *args):
        try:
            val = int(self.read_interval_us.get())
            self.read_interval_us.set(max(1000, min(200000, val)))
        except ValueError:
            self.read_interval_us.set(100000)

    def validate_auto_detect_threshold(self, *args):
        try:
            val = int(self.auto_detect_threshold.get())
            self.auto_detect_threshold.set(max(1, min(20, val)))
        except ValueError:
            self.auto_detect_threshold.set(3)

    def validate_alarm_interval(self, *args):
        try:
            val = int(self.alarm_interval_ms.get())
            self.alarm_interval_ms.set(max(10, min(2000, val)))
        except ValueError:
            self.alarm_interval_ms.set(100)

    def validate_alarm_sound_interval(self, *args):
        try:
            val = int(self.alarm_sound_interval_ms.get())
            self.alarm_sound_interval_ms.set(max(100, min(5000, val)))
        except ValueError:
            self.alarm_sound_interval_ms.set(500)

    def validate_tolerance(self, *args):
        try:
            val = float(self.tolerance_percent.get())
            self.tolerance_percent.set(max(0.0, min(100.0, val)))
        except ValueError:
            self.tolerance_percent.set(2.0)

    def validate_hold_time(self, *args):
        try:
            val = int(self.hold_time_ms.get())
            self.hold_time_ms.set(max(100, min(10000, val)))
        except ValueError:
            self.hold_time_ms.set(1000)

    def on_regex_selected(self, event=None):
        selected = self.data_format_regex_var.get()
        if selected == "Custom":
            custom = simpledialog.askstring("Custom Regex", "Enter regex to extract weight (use one capturing group):")
            if custom:
                try:
                    re.compile(custom)
                    self.data_format_regex_var.set(custom)
                    self.save_user_settings()
                except re.error as e:
                    messagebox.showerror("Invalid Regex", f"Invalid regex: {e}")
                    self.data_format_regex_var.set(r"(\d+\.?\d*)")
            else:
                self.data_format_regex_var.set(r"(\d+\.?\d*)")
        else:
            self.save_user_settings()

    def parse_weight_with_regex(self, line: str) -> float | None:
        pattern = self.data_format_regex_var.get()
        if not pattern:
            return None
        try:
            match = re.search(pattern, line)
            if match:
                return float(match.group(1))
        except (ValueError, IndexError, re.error):
            pass
        return None

    def _auto_detect_regex(self, line: str) -> str | None:
        for regex_pattern in self.predefined_regexes:
            if regex_pattern == "Custom":
                continue
            try:
                match = re.search(regex_pattern, line)
                if match and match.group(1):
                    float(match.group(1))
                    return regex_pattern
            except (ValueError, IndexError):
                continue
        return None

    def play_alarm_sound(self):
        if not os.path.isfile(ALARM_SOUND_FILE):
            logging.warning(f"Alarm sound file not found: {ALARM_SOUND_FILE}")
            return
        try:
            winsound.PlaySound(ALARM_SOUND_FILE, winsound.SND_ASYNC | winsound.SND_FILENAME)
        except Exception as e:
            logging.error(f"Failed to play alarm sound: {e}")

    def read_loop(self):
        self.regex_fail_count = 0
        if self.ser:
            self.ser.timeout = 1.0
        while not self.stop_read_event.is_set() and self.ser and self.ser.is_open:
            try:
                raw = self.ser.readline()
                if not raw and self.ser.in_waiting > 0:
                    raw = self.ser.read(self.ser.in_waiting)
                if raw:
                    try:
                        line = raw.decode('ascii', errors='ignore')
                    except:
                        line = raw.decode('latin1', errors='ignore')
                    self.last_serial_line = line
                    self.raw_serial_lines.append(line.strip())
                    val = self.parse_weight_with_regex(line)
                    if val is not None:
                        self.regex_fail_count = 0
                        decimals = self.decimal_places_var.get()
                        val_rounded = round(val, decimals)
                        if self.scale_units_var.get() == "kg":
                            weight_in_grams = val_rounded * 1000.0
                        else:
                            weight_in_grams = val_rounded
                        if self.lock_stable_var.get():
                            if self.stable_history.maxlen != self.stable_threshold_var.get():
                                self.stable_history = deque(maxlen=self.stable_threshold_var.get())
                            self.stable_history.append(weight_in_grams)
                            if (len(self.stable_history) == self.stable_history.maxlen and
                                len(set(self.stable_history)) == 1):
                                with self.weight_lock:
                                    self._current_weight = weight_in_grams
                        else:
                            with self.weight_lock:
                                self._current_weight = weight_in_grams
                        self.root.after(0, self.update_raw_serial_display)
                    else:
                        self.regex_fail_count += 1
                        if self.regex_fail_count >= self.auto_detect_threshold.get():
                            new_regex = self._auto_detect_regex(line)
                            if new_regex:
                                self.data_format_regex_var.set(new_regex)
                                self.save_user_settings()
                                self.log(f"Auto-detected data format: {new_regex}")
                            self.regex_fail_count = 0
                time.sleep(0.01)
            except Exception as e:
                self.log(f"Serial read error: {e}", level="ERROR")
                break

    def alarm_loop(self):
        last_rts_state = None
        last_sound_play_time = 0
        active_sound_status = None
        while not self.stop_alarm_event.is_set() and self.alarm_ser:
            try:
                rts_state = False
                current_status = "WAITING"
                if self.batch_active and self.alarm_connected:
                    current_status = self.current_status
                    rts_enabled = (
                        (current_status == "UNDER" and self.alarm_send_under.get()) or
                        (current_status == "OK" and self.alarm_send_ok.get()) or
                        (current_status == "OVER" and self.alarm_send_over.get())
                    )
                    rts_state = rts_enabled
                if rts_state != last_rts_state:
                    try:
                        self.alarm_ser.rts = rts_state
                        last_rts_state = rts_state
                        self.log(f"RTS {'ASSERTED' if rts_state else 'DEASSERTED'} for status: {current_status}")
                    except Exception as e:
                        self.log(f"Failed to set RTS: {e}", level="ERROR")
                now = time.time()
                sound_interval_sec = self.alarm_sound_interval_ms.get() / 1000.0
                desired_sound_status = None
                if current_status == "UNDER" and self.alarm_sound_under.get():
                    desired_sound_status = "UNDER"
                elif current_status == "OK" and self.alarm_sound_ok.get():
                    desired_sound_status = "OK"
                elif current_status == "OVER" and self.alarm_sound_over.get():
                    desired_sound_status = "OVER"
                should_play = False
                if desired_sound_status is not None:
                    if active_sound_status != desired_sound_status:
                        active_sound_status = desired_sound_status
                        last_sound_play_time = now
                        should_play = True
                    elif (now - last_sound_play_time) >= sound_interval_sec:
                        last_sound_play_time = now
                        should_play = True
                else:
                    if active_sound_status is not None:
                        active_sound_status = None
                if should_play:
                    self.play_alarm_sound()
                rts_interval_sec = max(0.01, self.alarm_interval_ms.get() / 1000.0)
                sleep_time = min(rts_interval_sec, sound_interval_sec / 2)
                time.sleep(sleep_time)
            except Exception as e:
                self.log(f"Alarm loop error: {e}", level="ERROR")
                break

    def update_raw_serial_display(self):
        if not hasattr(self, 'raw_display_var'):
            return
        display_text = "\n".join(self.raw_serial_lines) if self.raw_serial_lines else "[No data]"
        self.raw_display_var.set(display_text)

    def build_ui(self):
        menubar = tk.Menu(self.root)
        self.settings_menu = tk.Menu(menubar, tearoff=0)
        self.settings_menu.insert_command(0, label="Show Developer Panels", command=self.toggle_developer_panels)
        menubar.add_cascade(label="Settings", menu=self.settings_menu)
        self.root.config(menu=menubar)
        self.developer_frame = tk.Frame(self.main_container)
        top_container = tk.Frame(self.developer_frame)
        top_container.pack(fill=tk.X, padx=10, pady=5)
        top_container.columnconfigure(0, weight=1)
        top_container.columnconfigure(1, weight=1)

        # ------------------ INPUT SERIAL SETTINGS ------------------
        serial_frame = tk.LabelFrame(top_container, text="Input Serial", padx=5, pady=5)
        serial_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5), pady=5)
        self.connect_btn = tk.Button(serial_frame, text="Connect", bg="#2196F3", fg="white", width=8, command=self.connect_serial)
        self.connect_btn.grid(row=0, column=0, padx=2, pady=2, sticky="ew")
        self.disconnect_btn = tk.Button(serial_frame, text="Disconnect", bg="#9E9E9E", fg="white", width=8, command=self.disconnect_serial, state="disabled")
        self.disconnect_btn.grid(row=0, column=1, padx=2, pady=2, sticky="ew")
        ttk.Separator(serial_frame, orient=tk.HORIZONTAL).grid(row=1, column=0, columnspan=6, sticky="ew", pady=5)
        tk.Label(serial_frame, text="COM Port").grid(row=2, column=0, sticky="w", pady=2)
        self.com_combo = ttk.Combobox(serial_frame, textvariable=self.com_var, values=self.get_com_ports(), width=6)
        self.com_combo.grid(row=2, column=1, padx=2, sticky="ew")
        tk.Label(serial_frame, text="Baud").grid(row=2, column=2, sticky="w", pady=2)
        self.baud_combo = ttk.Combobox(serial_frame, textvariable=self.baud_var, values=["300","600","1200","2400","4800","9600","14400","19200","38400","57600","115200"], width=6)
        self.baud_combo.grid(row=2, column=3, padx=2, sticky="ew")
        tk.Label(serial_frame, text="Parity").grid(row=2, column=4, sticky="w", pady=2)
        self.parity_combo = ttk.Combobox(serial_frame, textvariable=self.parity_var, values=["None","Even","Odd"], state="readonly", width=6)
        self.parity_combo.grid(row=2, column=5, padx=2, sticky="ew")
        bottom_row = tk.Frame(serial_frame)
        bottom_row.grid(row=3, column=0, columnspan=6, sticky="ew", pady=2)
        tk.Label(bottom_row, text="Read Int (µs)").pack(side=tk.LEFT, padx=(0,2))
        read_interval_entry = ttk.Entry(bottom_row, textvariable=self.read_interval_us, width=6)
        read_interval_entry.pack(side=tk.LEFT, padx=(0,5))
        read_interval_entry.bind("<FocusOut>", self.validate_read_interval)
        tk.Label(bottom_row, text="Thresh").pack(side=tk.LEFT, padx=(10,2))
        threshold_entry = ttk.Entry(bottom_row, textvariable=self.auto_detect_threshold, width=4)
        threshold_entry.pack(side=tk.LEFT, padx=(0,5))
        threshold_entry.bind("<FocusOut>", self.validate_auto_detect_threshold)
        tk.Label(bottom_row, text="Scale Units").pack(side=tk.LEFT, padx=(10,2))
        scale_units_combo = ttk.Combobox(bottom_row, textvariable=self.scale_units_var, values=["g", "kg"], state="readonly", width=5)
        scale_units_combo.pack(side=tk.LEFT, padx=(0,5))
        tk.Label(bottom_row, text="Stable Lock").pack(side=tk.LEFT, padx=(10,2))
        lock_check = tk.Checkbutton(bottom_row, variable=self.lock_stable_var)
        lock_check.pack(side=tk.LEFT, padx=(0,2))
        tk.Label(bottom_row, text="Stable Thresh").pack(side=tk.LEFT, padx=(10,2))
        stable_thresh_combo = ttk.Combobox(
            bottom_row,
            textvariable=self.stable_threshold_var,
            values=[str(i) for i in range(1, 11)],
            width=3,
            state="readonly"
        )
        stable_thresh_combo.pack(side=tk.LEFT, padx=(0,5))
        tk.Label(bottom_row, text="Decimals").pack(side=tk.LEFT, padx=(10,2))
        decimals_combo = ttk.Combobox(
            bottom_row,
            textvariable=self.decimal_places_var,
            values=["0", "1", "2", "3", "4", "5"],
            width=3,
            state="readonly"
        )
        decimals_combo.pack(side=tk.LEFT, padx=(0,5))
        tk.Label(serial_frame, text="Data Format").grid(row=4, column=0, sticky="w", pady=2)
        self.data_format_regex_combo = ttk.Combobox(
            serial_frame,
            textvariable=self.data_format_regex_var,
            values=self.predefined_regexes,
            width=15,
            state="readonly"
        )
        self.data_format_regex_combo.grid(row=4, column=1, columnspan=5, padx=2, sticky="ew")
        self.data_format_regex_combo.bind("<<ComboboxSelected>>", self.on_regex_selected)

        # ------------------ ALARM SETTINGS ------------------
        alarm_frame = tk.LabelFrame(top_container, text="Alarm Output (RS232)", padx=5, pady=5)
        alarm_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0), pady=5)
        self.alarm_connect_btn = tk.Button(alarm_frame, text="Connect", bg="#2196F3", fg="white", width=8, command=self.connect_alarm_port)
        self.alarm_connect_btn.grid(row=0, column=0, padx=2, pady=2, sticky="ew")
        self.alarm_disconnect_btn = tk.Button(alarm_frame, text="Disconnect", bg="#9E9E9E", fg="white", width=8, command=self.disconnect_alarm_port, state="disabled")
        self.alarm_disconnect_btn.grid(row=0, column=1, padx=2, pady=2, sticky="ew")
        ttk.Separator(alarm_frame, orient=tk.HORIZONTAL).grid(row=1, column=0, columnspan=6, sticky="ew", pady=5)
        tk.Label(alarm_frame, text="COM Port").grid(row=2, column=0, sticky="w", pady=2)
        self.alarm_com_combo = ttk.Combobox(alarm_frame, textvariable=self.alarm_com_var, values=self.get_com_ports(), width=6)
        self.alarm_com_combo.grid(row=2, column=1, padx=2, sticky="ew")
        tk.Label(alarm_frame, text="Baud").grid(row=2, column=2, sticky="w", pady=2)
        self.alarm_baud_combo = ttk.Combobox(alarm_frame, textvariable=self.alarm_baud_var, values=["300","600","1200","2400","4800","9600","14400","19200","38400","57600","115200"], width=6)
        self.alarm_baud_combo.grid(row=2, column=3, padx=2, sticky="ew")
        tk.Label(alarm_frame, text="Parity").grid(row=2, column=4, sticky="w", pady=2)
        self.alarm_parity_combo = ttk.Combobox(alarm_frame, textvariable=self.alarm_parity_var, values=["None","Even","Odd"], state="readonly", width=6)
        self.alarm_parity_combo.grid(row=2, column=5, padx=2, sticky="ew")
        tk.Label(alarm_frame, text="Signal Types to Send:").grid(row=3, column=0, sticky="w", pady=(10, 2))
        under_check = tk.Checkbutton(alarm_frame, text="UNDER", variable=self.alarm_send_under)
        under_check.grid(row=3, column=1, sticky="w", padx=(0, 5))
        ok_check = tk.Checkbutton(alarm_frame, text="OK", variable=self.alarm_send_ok)
        ok_check.grid(row=3, column=2, sticky="w", padx=(0, 5))
        over_check = tk.Checkbutton(alarm_frame, text="OVER", variable=self.alarm_send_over)
        over_check.grid(row=3, column=3, sticky="w", padx=(0, 5))
        tk.Label(alarm_frame, text="Interval (ms):").grid(row=3, column=4, sticky="e", padx=(5, 0))
        interval_entry = ttk.Entry(alarm_frame, textvariable=self.alarm_interval_ms, width=6)
        interval_entry.grid(row=3, column=5, sticky="w", padx=(0, 5))
        tk.Label(alarm_frame, text="Alarm Sound:").grid(row=4, column=0, sticky="w", pady=(10, 2))
        sound_under_check = tk.Checkbutton(alarm_frame, text="UNDER", variable=self.alarm_sound_under)
        sound_under_check.grid(row=4, column=1, sticky="w", padx=(0, 5))
        sound_ok_check = tk.Checkbutton(alarm_frame, text="OK", variable=self.alarm_sound_ok)
        sound_ok_check.grid(row=4, column=2, sticky="w", padx=(0, 5))
        sound_over_check = tk.Checkbutton(alarm_frame, text="OVER", variable=self.alarm_sound_over)
        sound_over_check.grid(row=4, column=3, sticky="w", padx=(0, 5))
        tk.Label(alarm_frame, text="Interval (ms):").grid(row=4, column=4, sticky="e", padx=(5, 0))
        sound_interval_entry = ttk.Entry(alarm_frame, textvariable=self.alarm_sound_interval_ms, width=6)
        sound_interval_entry.grid(row=4, column=5, sticky="w", padx=(0, 5))
        sound_interval_entry.bind("<FocusOut>", self.validate_alarm_sound_interval)

        spacer = tk.Frame(top_container, height=10)
        spacer.grid(row=1, column=0, columnspan=2, sticky="ew", pady=2)

        # ------------------ SIMULATED WEIGHT (LEFT) ------------------
        sim_frame = tk.LabelFrame(top_container, text="Simulated Weight", padx=5, pady=3)
        sim_frame.grid(row=2, column=0, sticky="ew", padx=(0, 5), pady=(0, 5))
        sim_frame.columnconfigure(1, weight=1)
        simulate_check = tk.Checkbutton(sim_frame, text="Simulate Weight", variable=self.simulate_mode, command=self.on_sim_toggle)
        simulate_check.grid(row=0, column=0, columnspan=2, sticky="w", pady=1)
        tk.Label(sim_frame, text="Weight:").grid(row=1, column=0, sticky="w", pady=1)
        sim_scale = tk.Scale(sim_frame, from_=0, to=100000, orient=tk.HORIZONTAL, variable=self.sim_weight_var, length=200, sliderlength=15, width=12)
        sim_scale.grid(row=1, column=1, sticky="ew", padx=2, pady=1)
        sim_btns = tk.Frame(sim_frame)
        sim_btns.grid(row=2, column=0, columnspan=2, pady=(2, 1), sticky="w")
        for lbl, delta in [("+1k",1000),("+100",100),("+10",10),("0",None),("-10",-10),("-100",-100),("-1k",-1000)]:
            tk.Button(sim_btns, text=lbl, width=4, font=("Arial", 7), command=lambda d=delta: self.bump_sim(d)).pack(side=tk.LEFT, padx=1)

        # ------------------ PRINT TICKET SETTINGS (RIGHT) ------------------
        print_frame = tk.LabelFrame(top_container, text="Print Ticket", padx=5, pady=3)
        print_frame.grid(row=2, column=1, sticky="ew", padx=(5, 0), pady=(0, 5))
        print_frame.columnconfigure(1, weight=1)

        # SINGLE ROW: Auto-print + Mode Toggle + COM/Baud/Parity
        auto_print_row = tk.Frame(print_frame)
        auto_print_row.grid(row=0, column=0, columnspan=3, sticky="w", pady=1)

        print_check = tk.Checkbutton(
            auto_print_row,
            text="Auto-print after capture",
            variable=self.enable_print_var,
            command=self.on_print_toggle
        )
        print_check.pack(side=tk.LEFT, padx=(0, 8))

        # Printer mode toggle
        tk.Label(auto_print_row, text="Mode:", font=("Arial", 8)).pack(side=tk.LEFT)
        printer_mode_combo = ttk.Combobox(
            auto_print_row,
            textvariable=self.printer_mode_var,
            values=["COM", "DEFAULT"],
            width=8,
            font=("Arial", 8),
            state="readonly"
        )
        printer_mode_combo.pack(side=tk.LEFT, padx=(2, 6))

        # COM port controls frame (will be hidden/shown based on mode)
        self.com_controls_frame = tk.Frame(auto_print_row)
        self.com_controls_frame.pack(side=tk.LEFT, padx=(0, 0))

        tk.Label(self.com_controls_frame, text="COM:", font=("Arial", 8)).pack(side=tk.LEFT)
        printer_com_combo = ttk.Combobox(self.com_controls_frame, textvariable=self.printer_com_var, values=self.get_com_ports(), width=6, font=("Arial", 8))
        printer_com_combo.pack(side=tk.LEFT, padx=(2, 6))

        tk.Label(self.com_controls_frame, text="Baud:", font=("Arial", 8)).pack(side=tk.LEFT)
        printer_baud_combo = ttk.Combobox(
            self.com_controls_frame,
            textvariable=self.printer_baud_var,
            values=["1200","2400","4800","9600","19200","38400","57600","115200"],
            width=6,
            font=("Arial", 8),
            state="readonly"
        )
        printer_baud_combo.pack(side=tk.LEFT, padx=(2, 6))

        tk.Label(self.com_controls_frame, text="Parity:", font=("Arial", 8)).pack(side=tk.LEFT)
        printer_parity_combo = ttk.Combobox(
            self.com_controls_frame,
            textvariable=self.printer_parity_var,
            values=["None", "Even", "Odd"],
            width=5,
            font=("Arial", 8),
            state="readonly"
        )
        printer_parity_combo.pack(side=tk.LEFT, padx=(2, 0))

        # Bind printer mode changes to show/hide COM controls
        self.printer_mode_var.trace_add("write", lambda *a: self.on_printer_mode_change())

        # Header
        tk.Label(print_frame, text="Header:", font=("Arial", 8)).grid(row=1, column=0, sticky="w", pady=1)
        header_entry = tk.Entry(print_frame, textvariable=self.print_header_var, font=("Arial", 8))
        header_entry.grid(row=1, column=1, columnspan=2, sticky="ew", padx=(5,0), pady=1)

        # Paper width
        tk.Label(print_frame, text="Width (chars):", font=("Arial", 8)).grid(row=2, column=0, sticky="w", pady=1)
        width_spinbox = tk.Spinbox(
            print_frame,
            from_=24, to=48, increment=1,
            textvariable=self.print_paper_width,
            width=6,
            font=("Arial", 8)
        )
        width_spinbox.grid(row=2, column=1, sticky="w", padx=(5, 0), pady=1)

        # Printer status
        self.printer_status_label = tk.Label(
            print_frame,
            text="Printer: Disconnected",
            font=("Arial", 7),
            fg="#9E9E9E"
        )
        self.printer_status_label.grid(row=3, column=0, columnspan=2, sticky="w", pady=(2, 0))

        # Test Print button
        test_btn = tk.Button(
            print_frame,
            text="🖨️ Test Print",
            bg="#607D8B", fg="white",
            font=("Arial", 7, "bold"),
            command=self.test_print_ticket
        )
        test_btn.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(3, 0))

        # Auto-save trace
        for var in [self.enable_print_var, self.printer_com_var, self.printer_baud_var, self.printer_parity_var, self.print_header_var]:
            var.trace_add("write", lambda *a: self.save_user_settings())

        self.stable_history = deque(maxlen=self.stable_threshold_var.get())

        # ------------------ MAIN CONTENT FRAME ------------------
        self.main_frame = tk.Frame(self.main_container)
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        left = tk.Frame(self.main_frame)
        left.pack(side=tk.LEFT, fill=tk.Y, padx=(10, 5), pady=10)
        center = tk.Frame(self.main_frame)
        center.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 5), pady=10)
        right = tk.Frame(self.main_frame)
        right.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 10), pady=10)
        center.rowconfigure(0, weight=1)
        center.columnconfigure(0, weight=1)
        left.rowconfigure(0, weight=1)
        right.rowconfigure(0, weight=1)

        # Recipe Editor
        recipe_frame = tk.LabelFrame(left, text="Recipe Editor", padx=10, pady=10)
        recipe_frame.pack(fill=tk.BOTH, expand=True, padx=(10, 5), pady=10)
        top_row = tk.Frame(recipe_frame)
        top_row.grid(row=0, column=0, columnspan=4, sticky="ew", pady=(0, 5))
        for i in range(4):
            top_row.columnconfigure(i, weight=1)
        tk.Label(top_row, text="Recipe Name:").grid(row=0, column=0, sticky="w")
        self.recipe_name_var = tk.StringVar()
        self.recipe_combo = ttk.Combobox(top_row, textvariable=self.recipe_name_var, state="readonly", width=20)
        self.recipe_combo.grid(row=0, column=1, sticky="ew", padx=(0,5))
        self.recipe_combo.bind("<<ComboboxSelected>>", self.load_recipe)
        new_recipe_btn = tk.Button(top_row, text="New Recipe", bg="#2196F3", fg="white", command=self.create_new_recipe)
        new_recipe_btn.grid(row=0, column=2, padx=5, sticky="ew")
        load_btn = tk.Button(top_row, text="Load Recipe...", bg="#9C27B0", fg="white", command=self.load_recipe_from_file)
        load_btn.grid(row=0, column=3, padx=5, sticky="ew")

        batch_id_frame = tk.Frame(recipe_frame)
        batch_id_frame.grid(row=1, column=0, columnspan=4, sticky="ew", pady=(5, 5))
        for i in range(4):
            batch_id_frame.columnconfigure(i, weight=1)
        tk.Label(batch_id_frame, text="Batch ID (optional):").grid(row=0, column=0, sticky="w", padx=5)
        batch_id_entry = tk.Entry(batch_id_frame, textvariable=self.user_batch_id_var, width=20)
        batch_id_entry.grid(row=0, column=1, sticky="w", padx=5)
        lock_check = tk.Checkbutton(batch_id_frame, text="Lock Batch ID", variable=self.lock_batch_id)
        lock_check.grid(row=0, column=2, padx=5, sticky="w")

        resume_row = tk.Frame(recipe_frame)
        resume_row.grid(row=2, column=0, columnspan=4, sticky="ew", pady=(5, 5))
        for i in range(4):
            resume_row.columnconfigure(i, weight=1)
        tk.Label(resume_row, text="Resume Batch ID:").grid(row=0, column=0, sticky="w", padx=5)
        self.resume_batch_id_var = tk.StringVar()
        resume_entry = tk.Entry(resume_row, textvariable=self.resume_batch_id_var, width=20)
        resume_entry.grid(row=0, column=1, sticky="w", padx=5)
        resume_btn = tk.Button(resume_row, text="Resume Batch", bg="#4CAF50", fg="white", command=self.resume_batch_by_id)
        resume_btn.grid(row=0, column=2, padx=5, sticky="ew")

        tk.Label(recipe_frame, text="Ingredient").grid(row=3, column=0, sticky="w")
        tk.Label(recipe_frame, text="Target Weight (g)").grid(row=3, column=1, sticky="w")
        tk.Label(recipe_frame, text="Tol (%)").grid(row=3, column=2, sticky="w")
        self.ing_name_var = tk.StringVar()
        self.ing_weight_var = tk.StringVar()
        self.ing_tol_var = tk.StringVar()
        name_entry = ttk.Entry(recipe_frame, textvariable=self.ing_name_var, width=18)
        name_entry.grid(row=4, column=0, sticky="ew", padx=(0,5))
        weight_entry = ttk.Entry(recipe_frame, textvariable=self.ing_weight_var, width=12)
        weight_entry.grid(row=4, column=1, sticky="ew", padx=(0,5))
        tol_entry2 = ttk.Entry(recipe_frame, textvariable=self.ing_tol_var, width=8)
        tol_entry2.grid(row=4, column=2, sticky="w")
        add_btn = tk.Button(recipe_frame, text="Add Ingredient", bg="#4CAF50", fg="white", command=self.add_ingredient_and_save)
        add_btn.grid(row=4, column=3, sticky="ew", padx=5)

        btn_frame = tk.Frame(recipe_frame)
        btn_frame.grid(row=5, column=0, columnspan=4, sticky="ew", pady=5, padx=5)
        btn_frame.columnconfigure(0, weight=1)
        btn_frame.columnconfigure(1, weight=1)
        btn_frame.columnconfigure(2, weight=1)
        edit_btn = tk.Button(btn_frame, text="Edit Selected", bg="#FF9800", fg="white", command=self.edit_selected)
        edit_btn.grid(row=0, column=0, sticky="ew", padx=(0, 2))
        del_btn = tk.Button(btn_frame, text="Delete Selected", bg="#F44336", fg="white", command=self.delete_selected)
        del_btn.grid(row=0, column=1, sticky="ew", padx=(2, 2))
        clear_btn = tk.Button(btn_frame, text="Clear Recipe", bg="#9E9E9E", fg="white", command=self.clear_recipe)
        clear_btn.grid(row=0, column=2, sticky="ew", padx=(2, 0))

        recipe_frame.rowconfigure(6, weight=1)
        for i in range(4):
            recipe_frame.columnconfigure(i, weight=1)

        self.tree = ttk.Treeview(recipe_frame, columns=("name", "target", "tol", "actual", "status"), show="headings", height=14, selectmode="extended")
        self.tree.heading("name", text="Ingredient Name")
        self.tree.heading("target", text="Target (g)")
        self.tree.heading("tol", text="Tol (%)")
        self.tree.heading("actual", text="Actual")
        self.tree.heading("status", text="Status")
        self.tree.column("name", width=150, anchor="w")
        self.tree.column("target", width=100, anchor="e")
        self.tree.column("tol", width=80, anchor="e")
        self.tree.column("actual", width=110, anchor="e")
        self.tree.column("status", width=100, anchor="center")
        self.tree.grid(row=6, column=0, rowspan=8, columnspan=4, sticky="nsew", pady=(5,0))
        self.tree.tag_configure("UNDER", background="#FFF8E1")
        self.tree.tag_configure("OK", background="#E8F5E9")
        self.tree.tag_configure("OVER", background="#FFEBEE")
        self.tree.tag_configure("PENDING", background="#ECEFF1")
        self.tree.bind("<Double-1>", self.on_tree_double_click)

        # Live Weighing
        live_frame = tk.LabelFrame(center, text="Live Weighing", padx=10, pady=10)
        live_frame.pack(fill=tk.BOTH, expand=True)
        live_frame.grid_propagate(False)
        live_frame.columnconfigure(0, weight=1)
        live_frame.rowconfigure(0, weight=0)
        live_frame.rowconfigure(1, weight=0)
        live_frame.rowconfigure(2, weight=1)
        live_frame.rowconfigure(3, weight=0)
        live_frame.rowconfigure(4, weight=0)
        self.current_ing_label = tk.Label(live_frame, text="Current Ingredient: (none)", font=("Arial", 16, "bold"))
        self.current_ing_label.grid(row=0, column=0, sticky="w", padx=10, pady=(5,0))
        self.target_info_label = tk.Label(live_frame, text="Target: -     OK Range: -     Tol: -", font=("Arial", 12))
        self.target_info_label.grid(row=1, column=0, sticky="w", padx=10, pady=(5,0))
        self.weight_label = tk.Label(live_frame, text="0.00", font=("Consolas", 64, "bold"), fg=COLOR_TEXT_DEFAULT)
        self.weight_label.grid(row=2, column=0, sticky="nsew", padx=10, pady=10)
        self.status_label = tk.Label(live_frame, text="Status: WAITING", font=("Arial", 14))
        self.status_label.grid(row=3, column=0, sticky="w", padx=10, pady=(5,0))
        self.total_label = tk.Label(live_frame, text="Total Net: 0.00", font=("Arial", 14, "bold"))
        self.total_label.grid(row=4, column=0, sticky="e", padx=10, pady=(10,0))
        self.tx_count_label = tk.Label(live_frame, text="Transaction: -", font=("Arial", 12))
        self.tx_count_label.grid(row=5, column=0, sticky="e", padx=10, pady=(0,5))

        trans_view_btn = tk.Button(center, text="View Transaction History", bg="#9E9E9E", fg="white", font=("Arial", 11, "bold"),
                                  command=self.create_transaction_viewer)
        trans_view_btn.pack(fill=tk.X, padx=10, pady=5)

        # Weighing Controls
        weigh_frame = tk.LabelFrame(right, text="Weighing Controls", padx=10, pady=10)
        weigh_frame.pack(fill=tk.BOTH, expand=False)
        new_trans_btn = tk.Button(weigh_frame, text="New Transaction", bg="#2196F3", fg="white", font=("Arial", 11, "bold"), command=self.start_new_transaction)
        new_trans_btn.grid(row=0, column=0, columnspan=2, sticky="ew", pady=3)
        tk.Label(weigh_frame, text="Default Tolerance (%)").grid(row=1, column=0, sticky="w")
        tol_entry = ttk.Entry(weigh_frame, textvariable=self.tolerance_percent, width=12)
        tol_entry.grid(row=1, column=1, sticky="w", padx=5)
        self.tolerance_percent.trace_add("write", self.validate_tolerance)
        tk.Label(weigh_frame, text="Hold Time (ms)").grid(row=2, column=0, sticky="w")
        hold_entry = ttk.Entry(weigh_frame, textvariable=self.hold_time_ms, width=12)
        hold_entry.grid(row=2, column=1, sticky="w", padx=5)
        self.hold_time_ms.trace_add("write", self.validate_hold_time)
        auto_check = tk.Checkbutton(weigh_frame, text="Auto-advance when in tolerance", variable=self.auto_advance_enabled)
        auto_check.grid(row=3, column=0, columnspan=2, sticky="w", pady=(5,0))
        loop_check = tk.Checkbutton(weigh_frame, text="Auto-loop batch (restart after last)", variable=self.auto_loop_enabled)
        loop_check.grid(row=4, column=0, columnspan=2, sticky="w", pady=(5,0))
        tk.Label(weigh_frame, text="Display Units").grid(row=5, column=0, sticky="w")
        units_combo = ttk.Combobox(weigh_frame, values=["g", "kg", "lb"], textvariable=self.units, width=6, state="readonly")
        units_combo.grid(row=5, column=1, sticky="w", padx=5)
        ttk.Separator(weigh_frame, orient=tk.HORIZONTAL).grid(row=6, column=0, columnspan=2, sticky="ew", pady=8)
        self.start_btn = tk.Button(weigh_frame, text="Start Batch", bg="#4CAF50", fg="white", font=("Arial", 11, "bold"), command=self.start_batch)
        self.start_btn.grid(row=7, column=0, columnspan=2, sticky="ew", pady=3)
        self.capture_btn = tk.Button(weigh_frame, text="Capture Now", bg="#2196F3", fg="white", font=("Arial", 11, "bold"), command=self.capture_current)
        self.capture_btn.grid(row=8, column=0, columnspan=2, sticky="ew", pady=3)
        self.pause_btn = tk.Button(weigh_frame, text="Pause Batch", bg="#FF9800", fg="white", font=("Arial", 11, "bold"), command=self.pause_batch)
        self.pause_btn.grid(row=9, column=0, columnspan=2, sticky="ew", pady=3)
        self.tare_btn = tk.Button(weigh_frame, text="Tare Scale", bg="#9E9E9E", fg="white", font=("Arial", 11, "bold"), command=self.manual_tare)
        self.tare_btn.grid(row=10, column=0, columnspan=2, sticky="ew", pady=3)
        self.zero_btn = tk.Button(weigh_frame, text="Zero Scale", bg="#9E9E9E", fg="white", font=("Arial", 11, "bold"), command=self.manual_zero)
        self.zero_btn.grid(row=11, column=0, columnspan=2, sticky="ew", pady=3)
        self.next_btn = tk.Button(weigh_frame, text="Next Ingredient", bg="#607D8B", fg="white", font=("Arial", 11, "bold"), command=self.next_ingredient)
        self.next_btn.grid(row=12, column=0, columnspan=2, sticky="ew", pady=3)
        self.prev_btn = tk.Button(weigh_frame, text="Previous Ingredient", bg="#607D8B", fg="white", font=("Arial", 11, "bold"), command=self.previous_ingredient)
        self.prev_btn.grid(row=13, column=0, columnspan=2, sticky="ew", pady=3)

        # Raw Serial Monitor
        raw_monitor_frame = tk.LabelFrame(right, text="Raw Serial Monitor", padx=5, pady=5)
        raw_monitor_frame.pack(fill=tk.X, padx=(5, 10), pady=(10, 10))
        self.raw_display_var = tk.StringVar(value="[No data]")
        raw_display_label = tk.Label(
            raw_monitor_frame,
            textvariable=self.raw_display_var,
            font=("Consolas", 8),
            height=3,
            anchor="nw",
            justify=tk.LEFT,
            bg="#f8f8f8",
            relief="sunken",
            padx=5,
            pady=2
        )
        raw_display_label.pack(fill=tk.BOTH, expand=True)

        ports = self.get_com_ports()
        if ports:
            self.com_var.set(ports[0])
            self.alarm_com_var.set(ports[0])

    # =================== ALL REMAINING CORE METHODS ===================
    def get_default_com(self):
        ports = self.get_com_ports()
        return ports[0] if ports else "COM1"

    def get_com_ports(self):
        try:
            return [p.device for p in serial.tools.list_ports.comports()]
        except Exception:
            return []

    def start_new_transaction(self):
        if not self.current_recipe_name:
            messagebox.showinfo("Info", "No recipe is currently loaded.")
            return
        if not self.recipe:
            messagebox.showinfo("Info", "The current recipe is empty.")
            return
        if not messagebox.askyesno("Confirm New Transaction", f"Clear all recorded weights for '{self.current_recipe_name}'?"):
            return
        self.grand_total_net = 0.0
        self.batch_id = f"{self.current_recipe_name}_{datetime.now().strftime('%H%M%S')}"
        if self.start_btn:
            self.start_btn.configure(state="normal")
        for idx, item in enumerate(self.recipe):
            item["actual"] = None
            item["status"] = "PENDING"
        for idx, item in enumerate(self.recipe):
            tol_disp = "-" if item["tol"] is None else format_float(item["tol"])
            self.tree.item(
                self.tree.get_children()[idx],
                values=(item["name"], format_float(item["target"]), tol_disp, "-", "PENDING"),
                tags=("PENDING",)
            )
        self.net_total = 0.0
        self.update_total_label()
        self.update_current_labels()
        self.log(f"Started new transaction for recipe: {self.current_recipe_name} with Batch ID: {self.batch_id}")
        self.current_index = 0
        self.update_current_labels()
        self.highlight_current_ingredient()
        if self.current_recipe_name in self.batch_states:
            del self.batch_states[self.current_recipe_name]
        self.save_batch_states()
        self.root.after(100, self.update_current_labels)

    def _auto_loop_new_transaction(self):
        self.pause_batch()
        self.batch_id = self.get_consistent_batch_id()
        for idx, item in enumerate(self.recipe):
            item["actual"] = None
            item["status"] = "PENDING"
        for idx, item in enumerate(self.recipe):
            tol_disp = "-" if item["tol"] is None else format_float(item["tol"])
            self.tree.item(
                self.tree.get_children()[idx],
                values=(item["name"], format_float(item["target"]), tol_disp, "-", "PENDING"),
                tags=("PENDING",)
            )
        self.net_total = 0.0
        self.update_total_label()
        self.current_index = 0
        self.batch_active = True
        self.in_tolerance_since = None
        self.update_current_labels()
        self.highlight_current_ingredient()
        batch_state = self.batch_states.get(self.current_recipe_name, {})
        batch_state.update({
            "current_index": self.current_index,
            "actuals": [item["actual"] for item in self.recipe],
            "statuses": [item["status"] for item in self.recipe],
            "batch_id": self.batch_id
        })
        self.batch_states[self.current_recipe_name] = batch_state
        self.save_batch_states()
        self.log(f"Auto-loop: Started new transaction with Batch ID: {self.batch_id}")

    def start_batch(self):
        if not self.recipe:
            messagebox.showerror("No Recipe", "Please add ingredients to the recipe first.")
            return
        if self.batch_active:
            messagebox.showinfo("Info", "Batch already active.")
            return
        self.batch_id = self.get_consistent_batch_id()
        batch_state = self.batch_states.get(self.current_recipe_name, {})
        saved_index = batch_state.get("current_index", 0)
        saved_actuals = batch_state.get("actuals", [None] * len(self.recipe))
        saved_statuses = batch_state.get("statuses", ["PENDING"] * len(self.recipe))
        is_incomplete = any(status == "PENDING" for status in saved_statuses)
        if is_incomplete and self.current_recipe_name:
            self.log(f"Resuming incomplete batch {self.batch_id} for recipe '{self.current_recipe_name}'.")
        else:
            self.log(f"Starting new batch {self.batch_id} for recipe '{self.current_recipe_name}'.")
        if self.start_btn:
            self.start_btn.configure(state="disabled")
        self.batch_states[self.current_recipe_name] = {
            "current_index": saved_index,
            "actuals": saved_actuals,
            "statuses": saved_statuses,
            "batch_id": self.batch_id
        }
        self.net_total = 0.0
        self.update_total_label()
        self.batch_active = True
        self.current_index = saved_index
        self.in_tolerance_since = None
        self.update_current_labels()
        self.highlight_current_ingredient()

    def resume_batch_by_id(self):
        batch_id = self.resume_batch_id_var.get().strip()
        if not batch_id:
            messagebox.showwarning("Invalid Input", "Please enter a Batch ID to resume.")
            return
        recipe_name = None
        found_ingredient_name = None
        found_target_weight = None
        try:
            with open(CSV_FILE, "r", newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                headers = next(reader)
                for row in reader:
                    if row[0] == batch_id:
                        found_ingredient_name = row[4]
                        found_target_weight = float(row[5])
                        break
        except Exception as e:
            messagebox.showerror("Error", f"Could not search for batch ID in history: {e}")
            return
        if not found_ingredient_name or found_target_weight is None:
            messagebox.showerror("Error", f"No transactions found for Batch ID '{batch_id}'. Cannot resume.")
            return
        for name, recipe in self.recipes_db.items():
            for item in recipe:
                if item["name"] == found_ingredient_name and item["target"] == found_target_weight:
                    recipe_name = name
                    break
            if recipe_name:
                break
        if not recipe_name:
            messagebox.showerror("Error", f"Could not find a recipe containing ingredient '{found_ingredient_name}' with target weight {found_target_weight}g.")
            return
        self.current_recipe_name = recipe_name
        self.recipe = self.recipes_db[recipe_name].copy()
        actuals: list[float | None] = [None] * len(self.recipe)
        statuses = ["PENDING"] * len(self.recipe)
        try:
            with open(CSV_FILE, "r", newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                headers = next(reader)
                for row in reader:
                    if row[0] == batch_id:
                        idx = int(row[3]) - 1
                        if 0 <= idx < len(self.recipe):
                            actuals[idx] = float(row[9]) if row[9] else None
                            statuses[idx] = row[10] if row[10] else "PENDING"
        except Exception as e:
            messagebox.showerror("Error", f"Could not load transaction history for batch '{batch_id}': {e}")
            return
        for idx, item in enumerate(self.recipe):
            tol_disp = "-" if item["tol"] is None else format_float(item["tol"])
            actual_disp = "-" if actuals[idx] is None else format_float(actuals[idx])
            status_disp = statuses[idx]
            if idx < len(self.tree.get_children()):
                item_id = self.tree.get_children()[idx]
                self.tree.item(item_id, values=(item["name"], format_float(item["target"]), tol_disp, actual_disp, status_disp), tags=(status_disp,))
            else:
                self.tree.insert("", tk.END, values=(item["name"], format_float(item["target"]), tol_disp, actual_disp, status_disp), tags=(status_disp,))
        current_index = -1
        for idx in range(len(statuses)):
            if statuses[idx] == "PENDING":
                current_index = idx
                break
        self.net_total = sum([a for a in actuals if a is not None])
        self.batch_id = batch_id
        self.batch_states[self.current_recipe_name] = {
            "current_index": current_index,
            "actuals": actuals,
            "statuses": statuses,
            "batch_id": batch_id
        }
        self.update_current_labels()
        self.update_total_label()
        self.highlight_current_ingredient()
        self.log(f"Resumed batch '{batch_id}' for recipe '{recipe_name}'. Current ingredient index: {current_index}. Total Net: {format_float(self.net_total)}g")
        if self.start_btn is not None:
            self.start_btn.configure(state="normal")
        if self.trans_window and self.trans_window.winfo_exists():
            self.refresh_transaction_viewer()

    def delete_selected_transactions(self):
        if not self.trans_tree or not self.trans_tree.winfo_exists():
            return
        selected_items = self.trans_tree.selection()
        if not selected_items:
            return
        if not prompt_password_with_retry(self.root, "Password Required", "Enter password to delete transaction(s):"):
            return
        try:
            with open(CSV_FILE, "r", newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                headers = next(reader)
                all_rows = list(reader)
        except Exception as e:
            logging.warning(f"Could not read CSV for deletion: {e}")
            return
        selected_set = set()
        for item in selected_items:
            values = self.trans_tree.item(item, "values")
            if len(values) >= 3:
                selected_set.add((values[0], values[2]))
        rows_to_keep = []
        for row in all_rows:
            if len(row) >= 3:
                key = (row[0], row[2])
                if key not in selected_set:
                    rows_to_keep.append(row)
            else:
                rows_to_keep.append(row)
        try:
            with open(CSV_FILE, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                writer.writerows(rows_to_keep)
        except Exception as e:
            logging.error(f"Silent failure during CSV deletion write: {e}")
            return
        self.refresh_transaction_viewer()
        deleted_count = len(all_rows) - len(rows_to_keep)
        if deleted_count > 0:
            self.log(f"Deleted {deleted_count} transaction(s) after password authentication.")

    def connect_serial(self):
        if self.connected:
            return
        try:
            parity_map = {"None": serial.PARITY_NONE, "Even": serial.PARITY_EVEN, "Odd": serial.PARITY_ODD}
            stop_bits_map = {"1": serial.STOPBITS_ONE, "2": serial.STOPBITS_TWO}
            self.ser = serial.Serial(
                port=self.com_var.get(),
                baudrate=int(self.baud_var.get()),
                parity=parity_map.get(self.parity_var.get(), serial.PARITY_NONE),
                bytesize=int(self.data_bits_var.get()),
                stopbits=stop_bits_map.get(self.stop_bits_var.get(), serial.STOPBITS_ONE),
                timeout=1.0
            )
            self.connected = True
            self.stop_read_event.clear()
            self.read_thread = threading.Thread(target=self.read_loop, daemon=True)
            self.read_thread.start()
            if self.connect_btn:
                self.connect_btn.configure(state="disabled")
            if self.disconnect_btn:
                self.disconnect_btn.configure(state="normal")
            self.log(f"Connected to {self.ser.port} @ {self.ser.baudrate}")
        except Exception as e:
            self.connected = False
            self.ser = None
            messagebox.showerror("Serial Error", str(e))
            self.log(f"Serial connection failed: {e}", level="ERROR")

    def disconnect_serial(self):
        self.stop_read_event.set()
        if self.read_thread and self.read_thread.is_alive():
            self.read_thread.join(timeout=2)
        if self.ser:
            try:
                self.ser.close()
            except Exception:
                pass
        self.connected = False
        self.ser = None
        if self.connect_btn:
            self.connect_btn.configure(state="normal" if not self.simulate_mode.get() else "disabled")
        if self.disconnect_btn:
            self.disconnect_btn.configure(state="disabled")
        self.log("Disconnected serial.")

    def connect_alarm_port(self):
        if self.alarm_connected:
            return
        try:
            parity_map = {"None": serial.PARITY_NONE, "Even": serial.PARITY_EVEN, "Odd": serial.PARITY_ODD}
            stop_bits_map = {"1": serial.STOPBITS_ONE, "2": serial.STOPBITS_TWO}
            self.alarm_ser = serial.Serial(
                port=self.alarm_com_var.get(),
                baudrate=int(self.alarm_baud_var.get()),
                parity=parity_map.get(self.alarm_parity_var.get(), serial.PARITY_NONE),
                bytesize=int(self.alarm_data_bits_var.get()),
                stopbits=stop_bits_map.get(self.alarm_stop_bits_var.get(), serial.STOPBITS_ONE),
                timeout=0.2
            )
            self.alarm_connected = True
            self.stop_alarm_event.clear()
            self.alarm_read_thread = threading.Thread(target=self.alarm_loop, daemon=True)
            self.alarm_read_thread.start()
            if self.alarm_connect_btn:
                self.alarm_connect_btn.configure(state="disabled")
            if self.alarm_disconnect_btn:
                self.alarm_disconnect_btn.configure(state="normal")
            self.log(f"Connected Alarm Output to {self.alarm_ser.port} @ {self.alarm_ser.baudrate}")
        except Exception as e:
            self.alarm_connected = False
            self.alarm_ser = None
            messagebox.showerror("Alarm Port Error", str(e))
            self.log(f"Alarm port connection failed: {e}", level="ERROR")

    def disconnect_alarm_port(self):
        self.stop_alarm_event.set()
        if self.alarm_read_thread and self.alarm_read_thread.is_alive():
            self.alarm_read_thread.join(timeout=2)
        if self.alarm_ser:
            try:
                self.alarm_ser.close()
            except Exception:
                pass
        self.alarm_connected = False
        self.alarm_ser = None
        if self.alarm_connect_btn:
            self.alarm_connect_btn.configure(state="normal")
        if self.alarm_disconnect_btn:
            self.alarm_disconnect_btn.configure(state="disabled")
        self.log("Disconnected Alarm Output.")

    def grams_to_display(self, grams):
        unit = self.units.get()
        if unit == "g":
            return grams
        elif unit == "kg":
            return grams / 1000.0
        elif unit == "lb":
            return grams * 0.00220462262
        return grams

    def pause_batch(self):
        if not self.batch_active:
            return
        self.batch_active = False
        if self.current_recipe_name:
            self.batch_states[self.current_recipe_name] = {
                "current_index": self.current_index,
                "actuals": [item["actual"] for item in self.recipe],
                "statuses": [item["status"] for item in self.recipe],
                "batch_id": self.batch_id
            }
            self.save_batch_states()
        self.log("Batch paused.")
        self.update_current_labels()
        if self.check_for_batch_completion():
            self.show_batch_complete_popup()

    def next_ingredient(self):
        if not self.batch_active:
            return
        if self.current_index < len(self.recipe) - 1:
            self.current_index += 1
            if self.current_recipe_name:
                batch_state = self.batch_states.get(self.current_recipe_name, {})
                batch_state["current_index"] = self.current_index
                self.batch_states[self.current_recipe_name] = batch_state
                self.save_batch_states()  # Persist state
            self.in_tolerance_since = None
            self.update_current_labels()

    def previous_ingredient(self):
        if not self.batch_active:
            return
        if self.current_index > 0:
            self.current_index -= 1
            if self.current_recipe_name:
                batch_state = self.batch_states.get(self.current_recipe_name, {})
                batch_state["current_index"] = self.current_index
                self.batch_states[self.current_recipe_name] = batch_state
                self.save_batch_states()  # Persist state
            self.in_tolerance_since = None
            self.update_current_labels()

    def capture_current(self):
        if self.check_for_batch_completion() and not self.auto_loop_enabled.get():
            self.log("Batch is complete. Please click 'New Transaction' to continue.")
            return
        if not self.batch_active or self.current_index < 0 or self.current_index >= len(self.recipe):
            return
        item = self.recipe[self.current_index]
        actual = float(self.get_display_weight())
        tol_used = item["tol"] if item["tol"] is not None else self.tolerance_percent.get()
        target_disp = self.grams_to_display(item["target"])
        status, min_ok_disp, max_ok_disp = self.compute_status(actual, target_disp, tol_used)
        batch_state = self.batch_states.get(self.current_recipe_name, {})
        actuals = batch_state.get("actuals", [None] * len(self.recipe))
        statuses = batch_state.get("statuses", ["PENDING"] * len(self.recipe))
        actuals[self.current_index] = actual
        statuses[self.current_index] = status
        self.batch_states[self.current_recipe_name] = {
            "current_index": self.current_index,
            "actuals": actuals,
            "statuses": statuses,
            "batch_id": self.batch_id
        }
        tol_disp = "-" if item["tol"] is None else format_float(item["tol"])
        self.tree.item(
            self.tree.get_children()[self.current_index],
            values=(item["name"], format_float(item["target"]), tol_disp, format_float(actual), status),
            tags=(status,)
        )
        self.net_total = sum([a for a in actuals if a is not None])
        actual_g = actual * (1000 if self.units.get() == "kg" else 1)
        if self.units.get() == "lb":
            actual_g = actual / 0.00220462262
        if self.lock_batch_id.get() and self.auto_loop_enabled.get():
            self.grand_total_net += actual_g
        self.update_total_label()
        target_g = item["target"]
        dev_g = actual_g - target_g
        tx_count = 1
        if os.path.exists(CSV_FILE):
            try:
                with open(CSV_FILE, "r", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    next(reader)
                    tx_count = sum(1 for row in reader if row and row[0] == self.batch_id) + 1
            except Exception:
                tx_count = 1
        self.append_csv_record(
            transaction_number=tx_count,
            ingredient_index=self.current_index,
            ingredient_name=item["name"],
            target=target_g,
            tol=tol_used,
            min_ok=min_ok_disp * (1000 if self.units.get() == "kg" else 1) if self.units.get() != "lb" else min_ok_disp / 0.00220462262,
            max_ok=max_ok_disp * (1000 if self.units.get() == "kg" else 1) if self.units.get() != "lb" else max_ok_disp / 0.00220462262,
            actual=actual_g,
            status=status,
            deviation=dev_g,
            units=self.units.get()
        )
        # >>> PRINT TICKET <<<
        ticket_data = {
            "batch_id": self.batch_id,
            "timestamp": datetime.now().strftime("%m/%d/%Y %I:%M:%S %p"),
            "ingredient": item["name"],
            "target": item["target"],
            "actual": actual_g,
            "tolerance": tol_used,
            "status": status,
            "deviation": dev_g
        }
        self.print_transaction_ticket(ticket_data)
        self.log(f"[{item['name']}] target={format_float(target_disp)} {self.units.get()} "
                 f"actual={format_float(actual)} {self.units.get()} "
                 f"tol={format_float(tol_used)}% status={status} dev={format_float(dev_g)}g")
        if self.simulate_mode.get():
            self.sim_weight_var.set(0.0)
            self.log("Simulated weight zeroed.")
        elif self.connected and self.ser:
            self.send_tare_command()
        was_last = (self.current_index == len(self.recipe) - 1)
        if not was_last:
            self.current_index += 1
            if self.current_recipe_name:
                batch_state = self.batch_states.get(self.current_recipe_name, {})
                batch_state["current_index"] = self.current_index
                self.batch_states[self.current_recipe_name] = batch_state
            self.in_tolerance_since = None
            self.last_manual_capture = time.time()
            self.update_current_labels()
            self.log(f"Moved to next ingredient: {self.recipe[self.current_index]['name']}")
        else:
            self.log("Final ingredient captured.")
            if self.auto_loop_enabled.get():
                self.log("Auto-loop enabled: starting new transaction...")
                self.root.after(500, self._auto_loop_new_transaction)
            else:
                messagebox.showinfo(
                    "Batch Complete",
                    "You've reached the last ingredient.\n"
                    "To continue batching, please click 'New Transaction'."
                )
                self.batch_active = False
        self.save_batch_states()
        if self.trans_window and self.trans_window.winfo_exists():
            self.refresh_transaction_viewer()
        if self.check_for_batch_completion() and not self.auto_loop_enabled.get():
            self.show_batch_complete_popup()

    def compute_status(self, actual, target_disp, tol_percent):
        min_ok = target_disp * (1.0 - tol_percent/100.0)
        max_ok = target_disp * (1.0 + tol_percent/100.0)
        if actual < min_ok:
            return "UNDER", min_ok, max_ok
        elif actual > max_ok:
            return "OVER", min_ok, max_ok
        else:
            return "OK", min_ok, max_ok

    def get_display_weight(self):
        if self.simulate_mode.get():
            w = self.sim_weight_var.get()
        else:
            with self.weight_lock:
                w = self._current_weight
        return self.grams_to_display(w)

    def ui_tick(self):
        weight = self.get_display_weight()
        self.weight_label.config(text=format_float(weight, 2))
        if self.waiting_for_zero:
            if abs(weight) <= self.zero_tolerance:
                if self.zero_stable_since is None:
                    self.zero_stable_since = time.time()
                elif (time.time() - self.zero_stable_since) * 1000.0 >= 1000:
                    self.waiting_for_zero = False
                    self.zero_stable_since = None
                    self.log("Scale stabilized at zero. Ready for next ingredient.")
                    if self.current_index < len(self.recipe) - 1:
                        self.current_index += 1
                        if self.current_recipe_name:
                            batch_state = self.batch_states.get(self.current_recipe_name, {})
                            batch_state["current_index"] = self.current_index
                            self.batch_states[self.current_recipe_name] = batch_state
                        self.in_tolerance_since = None
                        self.last_manual_capture = time.time()
                        self.update_current_labels()
                        self.log(f"Moved to next ingredient: {self.recipe[self.current_index]['name']}")
                    else:
                        self.log("At last ingredient. Awaiting user action.")
                    self.save_batch_states()
            else:
                self.zero_stable_since = None
                self.status_label.config(text="Status: WAITING FOR ZERO", fg=COLOR_UNDER)
        else:
            status = "UNDER"
            if self.batch_active and (0 <= self.current_index < len(self.recipe)):
                item = self.recipe[self.current_index]
                tol_used = item["tol"] if item["tol"] is not None else self.tolerance_percent.get()
                target_disp = self.grams_to_display(item["target"])
                status, min_ok, max_ok = self.compute_status(weight, target_disp, tol_used)
                if status == "OK" and self.auto_advance_enabled.get():
                    now = time.time()
                    if self.in_tolerance_since is None:
                        self.in_tolerance_since = now
                    elif (now - self.in_tolerance_since) * 1000.0 >= self.hold_time_ms.get():
                        if now - self.last_manual_capture > 1.0:
                            if self.current_index < len(self.recipe) - 1:
                                self.capture_current()
                            else:
                                self.capture_current()
                                self.in_tolerance_since = None
            self.apply_status(status, only_visual=True)
        if self.batch_active and self.check_for_batch_completion() and not self.auto_loop_enabled.get():
            self.show_batch_complete_popup()
        self.update_current_labels()
        self.root.after(120, self.ui_tick)

    def apply_status(self, status, only_visual=False):
        if status != self.current_status:
            self.current_status = status
            self.start_flashing_for_status(status)
        color = COLOR_UNDER if status == "UNDER" else COLOR_OK if status == "OK" else COLOR_OVER
        self.status_label.config(text=f"Status: {status}", fg=color)

    def start_flashing_for_status(self, status):
        if self.flash_job:
            self.root.after_cancel(self.flash_job)
            self.flash_job = None
        self.flash_color = COLOR_UNDER if status == "UNDER" else COLOR_OK if status == "OK" else COLOR_OVER
        self.flash_on = False
        self.flash()

    def flash(self):
        self.flash_on = not self.flash_on
        fg = self.flash_color if self.flash_on else COLOR_TEXT_DEFAULT
        self.weight_label.config(fg=fg)
        self.flash_job = self.root.after(350, self.flash)

    def update_current_labels(self):
        if self.batch_active and 0 <= self.current_index < len(self.recipe):
            item = self.recipe[self.current_index]
            tol_used = item["tol"] if item["tol"] is not None else self.tolerance_percent.get()
            disp_target = self.grams_to_display(item["target"])
            disp_min = disp_target * (1 - tol_used/100.0)
            disp_max = disp_target * (1 + tol_used/100.0)
            self.current_ing_label.config(text=f"Current Ingredient: {item['name']} (#{self.current_index + 1})")
            self.target_info_label.config(text=f"Target: {format_float(disp_target)} {self.units.get()}    "
                                               f"OK Range: {format_float(disp_min)} - {format_float(disp_max)} {self.units.get()}    "
                                               f"Tol: {format_float(tol_used)}%")
            self.highlight_current_ingredient()
            tx_count = 1
            if os.path.exists(CSV_FILE):
                try:
                    with open(CSV_FILE, "r", encoding="utf-8") as f:
                        reader = csv.reader(f)
                        next(reader)
                        tx_count = sum(1 for row in reader if row and row[0] == self.batch_id) + 1
                except Exception:
                    tx_count = 1
            self.tx_count_label.config(text=f"Transaction: #{tx_count}")
        else:
            self.current_ing_label.config(text="Current Ingredient: (none)")
            self.target_info_label.config(text="Target: -     OK Range: -     Tol: -")
            self.tx_count_label.config(text="Transaction: -")

    def update_total_label(self):
        if self.lock_batch_id.get() and self.auto_loop_enabled.get():
            display_total = self.grams_to_display(self.grand_total_net)
            self.total_label.config(text=f"Total Net: {format_float(display_total)} {self.units.get()} (Accumulated)")
        else:
            self.total_label.config(text=f"Total Net: {format_float(self.net_total)} {self.units.get()}")

    def on_trans_window_close(self):
        if self.trans_window:
            self.trans_window.destroy()
            self.trans_window = None

    def load_all_transactions(self):
        if self.trans_tree is None:
            return
        for item in self.trans_tree.get_children():
            self.trans_tree.delete(item)
        try:
            with open(CSV_FILE, "r", newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                headers = next(reader)
                for row in reader:
                    row_to_display = row[:12]
                    self.trans_tree.insert("", tk.END, values=row_to_display)
        except Exception as e:
            self.log(f"Failed to load transactions: {e}", level="ERROR")

    def search_transactions(self):
        if self.trans_tree is None:
            return
        query = self.search_var.get().strip().lower()
        for item in self.trans_tree.get_children():
            self.trans_tree.delete(item)
        try:
            with open(CSV_FILE, "r", newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                headers = next(reader)
                for row in reader:
                    row_to_display = row[:12]
                    batch_id_match = query in row[0].lower()
                    date_match = query in row[2][:10].lower() if len(row[2]) >= 10 else False
                    ingredient_name_match = query in row[4].lower()
                    if batch_id_match or date_match or ingredient_name_match:
                        self.trans_tree.insert("", tk.END, values=row_to_display)
        except Exception as e:
            self.log(f"Search failed: {e}", level="ERROR")

    def clear_search(self):
        self.search_var.set("")
        self.load_all_transactions()

    def select_all_rows(self):
        if self.trans_tree is None:
            return
        children = self.trans_tree.get_children()
        if children:
            self.trans_tree.selection_set(children)

    def export_filtered_to_csv(self):
        if self.trans_tree is None:
            return
        filename = filedialog.asksaveasfilename(
            title="Export Filtered Transactions",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if not filename:
            return
        rows = []
        for item in self.trans_tree.get_children():
            values = self.trans_tree.item(item, "values")
            rows.append(values)
        try:
            with open(filename, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow([
                    "batch_id", "transaction_number", "timestamp", "ingredient_index", "ingredient_name",
                    "target_weight_g", "tolerance_percent", "min_ok_g", "max_ok_g",
                    "actual_weight_g", "status", "deviation_g"
                ])
                writer.writerows(rows)
            messagebox.showinfo("Success", f"Filtered transactions exported to {filename}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Could not export: {e}")

    def export_to_excel(self):
        filename = filedialog.asksaveasfilename(
            title="Export Transactions to Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not filename:
            return
        try:
            wb = Workbook()
            ws_list = wb.active
            if ws_list is None:
                messagebox.showerror("Export Error", "Could not create Excel workbook")
                return
            ws_list.title = "Transaction Batch List"
            headers = [
                "batch_id", "transaction_number", "timestamp", "ingredient_index", "ingredient_name",
                "target_weight_g", "tolerance_percent", "min_ok_g", "max_ok_g",
                "actual_weight_g", "status", "deviation_g"
            ]
            rows = []
            try:
                with open(CSV_FILE, "r", newline="", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    next(reader)
                    for row in reader:
                        rows.append(row[:12])
            except Exception as e:
                self.log(f"Failed to read CSV for export: {e}", level="ERROR")
                messagebox.showerror("Export Error", f"Could not read source: {e}")
                return
            for col_num, header in enumerate(headers, 1):
                cell = ws_list.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            for row_num, row_data in enumerate(rows, 2):
                for col_num, value in enumerate(row_data, 1):
                    ws_list.cell(row=row_num, column=col_num, value=value)
            for col in ws_list.columns:
                max_length = 0
                col_cell = col[0]
                if not hasattr(col_cell, 'column_letter'):
                    continue
                column: str = col_cell.column_letter  # type: ignore
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_list.column_dimensions[column].width = adjusted_width
            ws_summary = wb.create_sheet(title="Summary of Transaction")
            batch_summaries = {}
            for row in rows:
                batch_id = row[0]
                if batch_id not in batch_summaries:
                    batch_summaries[batch_id] = {
                        "start_time": row[2],
                        "end_time": row[2],
                        "total_net": 0.0,
                        "ingredient_count": 0,
                        "completed": True
                    }
                batch_summaries[batch_id]["end_time"] = row[2]
                try:
                    actual_weight = float(row[9])
                    batch_summaries[batch_id]["total_net"] += actual_weight
                except ValueError:
                    pass
                batch_summaries[batch_id]["ingredient_count"] += 1
                if not row[9]:
                    batch_summaries[batch_id]["completed"] = False
            summary_headers = ["Batch ID", "Start Time", "End Time", "Total Net Weight (g)", "Ingredient Count", "Status"]
            summary_data = []
            for batch_id, summary in batch_summaries.items():
                status = "Completed" if summary["completed"] else "Incomplete"
                summary_data.append([
                    batch_id,
                    summary["start_time"],
                    summary["end_time"],
                    summary["total_net"],
                    summary["ingredient_count"],
                    status
                ])
            for col_num, header in enumerate(summary_headers, 1):
                cell = ws_summary.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            for row_num, row_data in enumerate(summary_data, 2):
                for col_num, value in enumerate(row_data, 1):
                    ws_summary.cell(row=row_num, column=col_num, value=value)
            for col in ws_summary.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_summary.column_dimensions[column].width = adjusted_width
            ws_products = wb.create_sheet(title="Product Totals")
            product_totals = {}
            for row in rows:
                ingredient_name = row[4]
                try:
                    actual_weight = float(row[9])
                except ValueError:
                    actual_weight = 0.0
                if ingredient_name not in product_totals:
                    product_totals[ingredient_name] = 0.0
                product_totals[ingredient_name] = product_totals[ingredient_name] + actual_weight
            product_headers = ["Ingredient Name", "Total Weight (g)"]
            product_data = [[name, total] for name, total in product_totals.items()]
            for col_num, header in enumerate(product_headers, 1):
                cell = ws_products.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            for row_num, row_data in enumerate(product_data, 2):
                for col_num, value in enumerate(row_data, 1):
                    ws_products.cell(row=row_num, column=col_num, value=value)
            for col in ws_products.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_products.column_dimensions[column].width = adjusted_width
            wb.save(filename)
            messagebox.showinfo("Success", f"Data exported to Excel file: {filename}")
        except Exception as e:
            self.log(f"Excel export failed: {e}", level="ERROR")
            messagebox.showerror("Export Error", f"Could not export to Excel: {e}")

    def log(self, msg, level="INFO"):
        timestamp = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
        line = f"[{timestamp}] {level}: {msg}"
        print(line, end="\n")
        if level == "ERROR":
            logging.error(msg)
        elif level == "WARNING":
            logging.warning(msg)
        else:
            logging.info(msg)

    def append_csv_record(self, transaction_number, ingredient_index, ingredient_name, target, tol, min_ok, max_ok, actual, status, deviation, units):
        ensure_csv_header(CSV_FILE)
        with open(CSV_FILE, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                self.batch_id or "",
                transaction_number,
                datetime.now().strftime("%m/%d/%Y %I:%M:%S %p"),
                ingredient_index + 1,
                ingredient_name,
                format_float(target),
                format_float(tol),
                format_float(min_ok),
                format_float(max_ok),
                format_float(actual),
                status,
                format_float(deviation),
                units
            ])

    def bump_sim(self, delta):
        if delta is None:
            self.sim_weight_var.set(0.0)
        else:
            self.sim_weight_var.set(max(0.0, self.sim_weight_var.get() + delta))

    def open_csv_file(self):
        path = os.path.abspath(CSV_FILE)
        if sys.platform == "win32":
            os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.run(["open", path])
        else:
            subprocess.run(["xdg-open", path])

    def load_batch_states(self):
        if os.path.exists(BATCH_STATES_FILE):
            try:
                with open(BATCH_STATES_FILE, "r", encoding="utf-8") as f:
                    self.batch_states = json.load(f)
                self.log("Batch states loaded.")
            except Exception as e:
                self.log(f"Failed to load batch states: {e}", level="ERROR")
                self.batch_states = {}
        else:
            self.batch_states = {}

    def save_batch_states(self):
        try:
            with open(BATCH_STATES_FILE, "w", encoding="utf-8") as f:
                json.dump(self.batch_states, f, indent=2)
            self.log("Batch states saved.")
        except Exception as e:
            messagebox.showerror("Save Error", f"Could not save batch states: {e}")

    def on_close(self):
        try:
            # Save all persistent state before closing
            self.save_user_settings()
            self.save_recipes_db()
            self.save_batch_states()
            self.disconnect_serial()
            self.disconnect_alarm_port()
            self.disconnect_printer()
        except Exception:
            pass
        self.root.destroy()

    def send_tare_command(self):
        if not self.connected or self.ser is None:
            self.log("Cannot send Tare command: No active serial connection.", level="WARNING")
            return
        try:
            command_str = "T"
            payload_bytes = bytes(ord(c) for c in command_str.upper()) + b'\r'
            bytes_written = self.ser.write(payload_bytes)
            self.ser.flush()
            self.log(f"Sent Tare command: '{command_str}' bytes {list(payload_bytes)} ({bytes_written} bytes)")
        except Exception as e:
            self.log(f"Failed to send Tare command: {e}", level="ERROR")

    def manual_tare(self):
        if not self.connected:
            messagebox.showwarning("No Connection", "Cannot send Tare command: No active serial connection.")
            return
        try:
            self.send_tare_command()
        except Exception as e:
            messagebox.showerror("Tare Error", f"Failed to send Tare command: {e}")

    def manual_zero(self):
        if not self.connected or self.ser is None:
            messagebox.showwarning("No Connection", "Cannot send Zero command: No active serial connection.")
            return
        try:
            command_str = "Z"
            payload_bytes = bytes(ord(c) for c in command_str.upper()) + b'\r'
            bytes_written = self.ser.write(payload_bytes)
            self.ser.flush()
            self.log(f"Sent Zero command: '{command_str}' bytes {list(payload_bytes)} ({bytes_written} bytes)")
        except Exception as e:
            messagebox.showerror("Zero Error", f"Failed to send Zero command: {e}")

    def highlight_current_ingredient(self):
        self.tree.selection_remove(self.tree.selection())
        if 0 <= self.current_index < len(self.recipe):
            item_id = self.tree.get_children()[self.current_index]
            self.tree.selection_set(item_id)
            self.tree.see(item_id)

    def copy_batch_id(self):
        if self.trans_tree is None:
            return
        selected_items = self.trans_tree.selection()
        if not selected_items:
            return
        item = selected_items[0]
        values = self.trans_tree.item(item, "values")
        batch_id = values[0]
        self.root.clipboard_clear()
        self.root.clipboard_append(batch_id)
        self.log(f"Batch ID '{batch_id}' copied to clipboard.")

    def refresh_transaction_viewer(self):
        if self.trans_window is None or not self.trans_window.winfo_exists() or self.trans_tree is None:
            return
        selected_items = self.trans_tree.selection()
        self.load_all_transactions()
        if selected_items:
            self.trans_tree.selection_set(selected_items)
            self.trans_tree.focus(selected_items[0] if selected_items else "")

    def check_for_batch_completion(self):
        if not self.batch_active or not self.recipe or not self.batch_id:
            return False
        for item in self.recipe:
            if item["actual"] is None:
                return False
        return True

    def show_batch_complete_popup(self):
        self.pause_batch()
        messagebox.showinfo(
            "Batch Complete!",
            "All ingredients for this recipe have been successfully recorded.\n"
            "To start a new batch, please click the 'New Transaction' button in the Recipe Editor."
        )

    def create_transaction_viewer(self):
        if self.trans_window and self.trans_window.winfo_exists():
            self.trans_window.lift()
            self.trans_window.focus_force()
            return
        self.trans_window = tk.Toplevel(self.root)
        self.trans_window.title("Transaction History Viewer")
        screen_width = self.trans_window.winfo_screenwidth()
        screen_height = self.trans_window.winfo_screenheight()
        self.trans_window.geometry(f"{screen_width}x{screen_height}+0+0")
        self.trans_window.state('zoomed')
        self.trans_window.protocol("WM_DELETE_WINDOW", self.on_trans_window_close)
        main_frame = tk.Frame(self.trans_window)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        main_frame.grid_propagate(False)
        main_frame.rowconfigure(0, weight=0)
        main_frame.rowconfigure(1, weight=1)
        main_frame.rowconfigure(2, weight=0)
        main_frame.columnconfigure(0, weight=1)
        filter_frame = tk.LabelFrame(main_frame, text="Search & Filter", padx=10, pady=10)
        filter_frame.grid(row=0, column=0, sticky="ew", padx=0, pady=5)
        filter_frame.columnconfigure(0, weight=0)
        filter_frame.columnconfigure(1, weight=1)
        filter_frame.columnconfigure(2, weight=1)
        filter_frame.columnconfigure(3, weight=1)
        tk.Label(filter_frame, text="Search Query:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.search_var = tk.StringVar()
        search_entry = tk.Entry(filter_frame, textvariable=self.search_var, width=30)
        search_entry.grid(row=0, column=1, sticky="ew", padx=5, pady=5)
        search_btn = tk.Button(filter_frame, text="Search", bg="#4CAF50", fg="white", command=self.search_transactions)
        search_btn.grid(row=0, column=2, sticky="ew", padx=5, pady=5)
        clear_btn = tk.Button(filter_frame, text="Clear Search", bg="#9E9E9E", fg="white", command=self.clear_search)
        clear_btn.grid(row=0, column=3, sticky="ew", padx=5, pady=5)
        tree_frame = tk.Frame(main_frame)
        tree_frame.grid(row=1, column=0, sticky="nsew", padx=0, pady=5)
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)
        columns = ["batch_id", "transaction_number", "timestamp", "ingredient_index", "ingredient_name",
                   "target_weight_g", "tolerance_percent", "min_ok_g", "max_ok_g",
                   "actual_weight_g", "status", "deviation_g"]
        self.trans_tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
        for col in columns:
            self.trans_tree.heading(col, text=col.replace("_", " ").title(), command=lambda c=col: self.sort_transactions(c))
            if col == "batch_id":
                self.trans_tree.column(col, width=150, anchor="center")
            elif col == "timestamp":
                self.trans_tree.column(col, width=180, anchor="center")
            elif col == "ingredient_name":
                self.trans_tree.column(col, width=200, anchor="center")
            elif col in ["target_weight_g", "actual_weight_g"]:
                self.trans_tree.column(col, width=100, anchor="center")
            else:
                self.trans_tree.column(col, width=80, anchor="center")
        self.trans_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.context_menu = tk.Menu(self.trans_tree, tearoff=0)
        self.context_menu.add_command(label="Copy Batch ID", command=self.copy_batch_id)
        def show_context_menu(event):
            if self.trans_tree is not None:
                item = self.trans_tree.identify_row(event.y)
                if item:
                    self.trans_tree.selection_set(item)
                    self.trans_tree.focus(item)
                    self.context_menu.post(event.x_root, event.y_root)
        self.trans_tree.bind("<Button-3>", show_context_menu)
        action_frame = tk.Frame(main_frame)
        action_frame.grid(row=2, column=0, sticky="ew", padx=0, pady=5)
        for i in range(4):
            action_frame.columnconfigure(i, weight=1)
        select_all_btn = tk.Button(action_frame, text="Select All", bg="#4CAF50", fg="white", font=("Arial", 11, "bold"), command=self.select_all_rows)
        select_all_btn.grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        delete_btn = tk.Button(action_frame, text="Delete", bg="#F44336", fg="white", font=("Arial", 11, "bold"), command=self.delete_selected_transactions)
        delete_btn.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        export_filtered_btn = tk.Button(action_frame, text="EXPORT FILTERED TO CSV", bg="#4CAF50", fg="white", font=("Arial", 11, "bold"), command=self.export_filtered_to_csv)
        export_filtered_btn.grid(row=0, column=2, padx=5, pady=5, sticky="ew")
        export_excel_btn = tk.Button(action_frame, text="EXPORT TO EXCEL", bg="#2196F3", fg="white", font=("Arial", 11, "bold"), command=self.export_to_excel)
        export_excel_btn.grid(row=0, column=3, padx=5, pady=5, sticky="ew")
        self.load_all_transactions()
        self.refresh_transaction_viewer()

    def sort_transactions(self, col):
        if self.trans_tree is None:
            return
        data = [(self.trans_tree.set(child, col), child) for child in self.trans_tree.get_children()]
        data.sort()
        for index, (val, child) in enumerate(data):
            self.trans_tree.move(child, "", index)

    def load_recipes_db(self):
        if os.path.exists(RECIPES_DB_FILE):
            try:
                with open(RECIPES_DB_FILE, "r", encoding="utf-8") as f:
                    self.recipes_db = json.load(f)
                if self.recipe_combo is not None:
                    self.recipe_combo['values'] = list(self.recipes_db.keys())
            except Exception:
                self.recipes_db = {}
        else:
            self.recipes_db = {}

    def save_recipes_db(self):
        try:
            with open(RECIPES_DB_FILE, "w", encoding="utf-8") as f:
                json.dump(self.recipes_db, f, indent=2)
            self.log(f"Recipes database saved to {RECIPES_DB_FILE}")
        except Exception as e:
            messagebox.showerror("Save Error", f"Could not save recipes: {e}")

    def create_new_recipe(self):
        name = simpledialog.askstring("New Recipe", "Enter recipe name:")
        if not name:
            return
        name = name.strip()
        if not name:
            return
        if name in self.recipes_db:
            if not messagebox.askyesno("Confirm", f"Recipe '{name}' already exists. Overwrite?"):
                return
        # ✅ Validate name
        if not name.strip():
            messagebox.showerror("Invalid Name", "Recipe name cannot be empty.")
            return
        self.recipes_db[name] = []
        self.current_recipe_name = name
        self.recipe = []
        self.recipe_name_var.set(name)
        if self.recipe_combo is not None:
            self.recipe_combo['values'] = list(self.recipes_db.keys())
            self.recipe_combo.set(name)
        if self.tree is not None:
            for i in self.tree.get_children():
                self.tree.delete(i)
        self.update_current_labels()
        self.log(f"Created new recipe: {name}")

    def load_recipe_from_file(self):
        filepath = filedialog.askopenfilename(
            title="Load Recipe from JSON",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        if not filepath:
            return
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
            # Assume root is recipe list or dict
            if isinstance(data, list):
                loaded_recipe = data
                recipe_name = os.path.splitext(os.path.basename(filepath))[0]
            elif isinstance(data, dict):
                if "recipe" in data and "name" in data:
                    loaded_recipe = data["recipe"]
                    recipe_name = data["name"]
                else:
                    # Top-level keys are recipe names
                    if len(data) == 1:
                        recipe_name, loaded_recipe = next(iter(data.items()))
                    else:
                        messagebox.showerror("Invalid Format", "JSON must contain exactly one recipe.")
                        return
            else:
                messagebox.showerror("Invalid Format", "Unsupported JSON structure.")
                return
            # Validate
            for ing in loaded_recipe:
                if not all(k in ing for k in ["name", "target"]):
                    messagebox.showerror("Invalid Format", "Each ingredient must have 'name' and 'target'.")
                    return
            self.recipes_db[recipe_name] = loaded_recipe
            self.save_recipes_db()
            if self.recipe_combo is not None:
                self.recipe_combo['values'] = list(self.recipes_db.keys())
            self.recipe_name_var.set(recipe_name)
            self.load_recipe()
            self.log(f"Loaded recipe from file: {filepath}")
        except Exception as e:
            messagebox.showerror("Load Error", f"Failed to load recipe: {e}")

    def load_recipe(self, event=None):
        if self.batch_active:
            self.pause_batch()
        name = self.recipe_name_var.get().strip()
        if not name:
            messagebox.showwarning("No Recipe Selected", "Please select a recipe from the dropdown.")
            return
        if name not in self.recipes_db:
            messagebox.showerror("Recipe Not Found", f"Recipe '{name}' does not exist in the database.")
            return
        self.current_recipe_name = name
        self.recipe = self.recipes_db[name].copy()
        batch_state = self.batch_states.get(name, {})
        saved_index = batch_state.get("current_index", 0)
        saved_actuals = batch_state.get("actuals", [None] * len(self.recipe))
        saved_statuses = batch_state.get("statuses", ["PENDING"] * len(self.recipe))
        while len(saved_actuals) < len(self.recipe):
            saved_actuals.append(None)
        while len(saved_statuses) < len(self.recipe):
            saved_statuses.append("PENDING")
        current_batch_id = batch_state.get("batch_id", None)
        if current_batch_id:
            try:
                with open(CSV_FILE, "r", newline="", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    headers = next(reader)
                    for row in reader:
                        if row[0] == current_batch_id:
                            idx = int(row[3]) - 1
                            if idx < len(self.recipe):
                                self.recipe[idx]["actual"] = float(row[9]) if row[9] else None
                                self.recipe[idx]["status"] = row[10] if row[10] else "PENDING"
                                saved_actuals[idx] = self.recipe[idx]["actual"]
                                saved_statuses[idx] = self.recipe[idx]["status"]
            except Exception as e:
                self.log(f"Failed to synchronize with CSV for batch {current_batch_id}: {e}", level="ERROR")
        for i in self.tree.get_children():
            self.tree.delete(i)
        for idx, item in enumerate(self.recipe):
            tol_disp = "-" if item["tol"] is None else format_float(item["tol"])
            actual_disp = "-" if item["actual"] is None else format_float(item["actual"])
            status_disp = item["status"]
            self.tree.insert("", tk.END, values=(item["name"], format_float(item["target"]), tol_disp, actual_disp, status_disp),
                             tags=(status_disp,))
        self.current_index = 0
        self.net_total = sum([a for a in saved_actuals if a is not None])
        pending_found = False
        for idx, item in enumerate(self.recipe):
            if item["status"] == "PENDING":
                self.current_index = idx
                pending_found = True
                break
        self.update_current_labels()
        self.update_total_label()
        self.log(f"Loaded recipe: {name} (resumed from index {self.current_index})")

    def add_ingredient(self):
        if not self.current_recipe_name:
            messagebox.showinfo("Info", "Please create or load a recipe first.")
            return
        name = self.ing_name_var.get().strip()
        if not name:
            messagebox.showerror("Invalid Input", "Ingredient name is required.")
            return
        # ✅ Check for duplicates
        name_lower = name.lower()
        for item in self.recipe:
            if item["name"].lower() == name_lower:
                messagebox.showerror("Duplicate Ingredient", "An ingredient with this name already exists.")
                return
        try:
            target = float(self.ing_weight_var.get())
        except Exception:
            messagebox.showerror("Invalid Input", "Target weight must be a number (grams).")
            return
        tol_txt = self.ing_tol_var.get().strip()
        tol_val = None
        if tol_txt != "":
            try:
                tol_val = float(tol_txt)
                if tol_val < 0:
                    raise ValueError
            except Exception:
                messagebox.showerror("Invalid Input", "Tolerance must be a non-negative number (or leave blank to use default).")
                return
        item = {"name": name, "target": target, "tol": tol_val, "actual": None, "status": "PENDING"}
        self.recipe.append(item)
        tol_disp = "-" if tol_val is None else format_float(tol_val)
        self.tree.insert("", tk.END, values=(name, format_float(target), tol_disp, "-", "PENDING"), tags=("PENDING",))
        self.ing_name_var.set("")
        self.ing_weight_var.set("")
        self.ing_tol_var.set("")
        self.update_current_labels()

    def add_ingredient_and_save(self):
        self.add_ingredient()
        self.save_recipes_db()

    def delete_selected(self):
        sel = self.tree.selection()
        if not sel:
            return
        if not messagebox.askyesno("Confirm Deletion", "Are you sure you want to permanently delete the selected ingredient(s) from the database?"):
            return
        idxs = sorted([self.tree.index(i) for i in sel], reverse=True)
        for idx in idxs:
            self.tree.delete(self.tree.get_children()[idx])
            del self.recipe[idx]
        if self.current_recipe_name and self.current_recipe_name in self.recipes_db:
            self.recipes_db[self.current_recipe_name] = self.recipe.copy()
            self.save_recipes_db()
            self.log(f"Deleted selected ingredients from recipe: {self.current_recipe_name}")
        self.update_current_labels()

    def clear_recipe(self):
        if not self.current_recipe_name:
            return
        if messagebox.askyesno("Confirm", "Clear the entire recipe?"):
            for i in self.tree.get_children():
                self.tree.delete(i)
            self.recipe.clear()
            if self.current_recipe_name in self.recipes_db:
                self.recipes_db[self.current_recipe_name] = []
                self.save_recipes_db()  # Persist cleared recipe
            self.update_current_labels()
            if self.start_btn:
                self.start_btn.configure(state="normal")

# ----------------------------
# Main Function
# ----------------------------
def main():
    root = tk.Tk()
    app = BatchingApp(root)
    root.protocol("WM_DELETE_WINDOW", app.on_close)
    root.state('zoomed')
    root.mainloop()

if __name__ == "__main__":
    main()